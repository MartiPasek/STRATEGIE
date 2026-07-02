"""
Ceníkový engine (Marti 2.7.2026) — port EUROSOFT DB-Ceniky do STRATEGIE.

Bezpečný evaluátor výrazových vzorců (Martiho @P styl z EC_CenikyVzorce), BEZ
dynamického SQL. Podporuje: @Pnn parametry, řetězcové literály '...', čísla,
operátory + - * /, a funkce SUBSTRING/LEFT/RIGHT/REPLACE/CAST/CHAR/UPPER/LOWER/
LTRIM/RTRIM/LEN/ISNULL. '+' = concat (řetězce) nebo součet (čísla). SQL 1-indexace
u SUBSTRING.

transform_row(vzorce, params) aplikuje uspořádané vzorce: cíl může být výstupní
pole (RegCisHeo, EC_NC…) NEBO pracovní @Pnn slot (přepíše se v params pro další vzorce).
"""
from __future__ import annotations

import re
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP

# ── tokenizer ────────────────────────────────────────────────────────────────
_TOKEN_RE = re.compile(r"""
    \s+
  | (?P<STR>'(?:[^']|'')*')
  | (?P<NUM>\d+\.\d+|\d+)
  | (?P<PARAM>@P\d{1,2}|@[A-Za-z_]\w*)
  | (?P<IDENT>[A-Za-z_]\w*)
  | (?P<CMP><=|>=|<>|!=|=|<|>)
  | (?P<OP>[+\-*/(),])
""", re.VERBOSE)


def _tokenize(s: str):
    toks = []
    i = 0
    while i < len(s):
        m = _TOKEN_RE.match(s, i)
        if not m or m.end() == i:
            raise ValueError("neplatný znak ve vzorci u: %r" % s[i:i + 12])
        i = m.end()
        if m.lastgroup is None:
            continue  # whitespace
        toks.append((m.lastgroup, m.group()))
    return toks


# ── parser (recursive descent) → AST ─────────────────────────────────────────
class _P:
    def __init__(self, toks):
        self.t = toks
        self.i = 0

    def peek(self):
        return self.t[self.i] if self.i < len(self.t) else (None, None)

    def next(self):
        tok = self.peek()
        self.i += 1
        return tok

    def expect(self, val):
        k, v = self.next()
        if v is None or v.upper() != val.upper():
            raise ValueError("čekal jsem %r, dostal %r" % (val, v))

    def parse(self):
        node = self.expr()
        if self.i != len(self.t):
            raise ValueError("nadbytečné tokeny ve vzorci")
        return node

    def expr(self):  # + -
        node = self.term()
        while self.peek()[1] in ("+", "-"):
            op = self.next()[1]
            node = ("bin", op, node, self.term())
        return node

    def term(self):  # * /
        node = self.factor()
        while self.peek()[1] in ("*", "/"):
            op = self.next()[1]
            node = ("bin", op, node, self.factor())
        return node

    def factor(self):
        k, v = self.peek()
        if v == "(":
            self.next(); node = self.expr(); self.expect(")"); return node
        if k == "STR":
            self.next(); return ("str", v[1:-1].replace("''", "'"))
        if k == "NUM":
            self.next(); return ("num", v)
        if k == "PARAM":
            self.next(); return ("param", v[1:])  # bez @
        if k == "IDENT":
            up = v.upper()
            if up == "NULL":
                self.next(); return ("null",)
            if up == "CASE":
                self.next(); return self.case_expr()
            self.next()
            if self.peek()[1] == "(":
                return self.func_call(v)
            # holé ident (např. typ v CAST) — vrátíme jako řetězec-název
            return ("name", v)
        raise ValueError("neočekávaný token: %r" % (v,))

    def _kw(self):
        v = self.peek()[1]
        return str(v).upper() if v is not None else ""

    def _expect_kw(self, kw):
        k, v = self.next()
        if not v or str(v).upper() != kw:
            raise ValueError("čekal jsem %s, dostal %r" % (kw, v))

    def condition(self):
        """Porovnání (a > b, a = b, …) nebo prostý výraz."""
        left = self.expr()
        if self.peek()[0] == "CMP":
            op = self.next()[1]
            return ("cmp", op, left, self.expr())
        return left

    def case_expr(self):
        """CASE [operand] WHEN … THEN … [ELSE …] END (searched i simple)."""
        searched = (self._kw() == "WHEN")
        operand = None if searched else self.expr()
        branches = []
        while self._kw() == "WHEN":
            self.next()
            whenpart = self.condition() if searched else self.expr()
            self._expect_kw("THEN")
            branches.append((whenpart, self.expr()))
        elseval = None
        if self._kw() == "ELSE":
            self.next(); elseval = self.expr()
        self._expect_kw("END")
        return ("case", searched, operand, branches, elseval)

    def func_call(self, name):
        self.expect("(")
        args = []
        if self.peek()[1] != ")":
            args.append(self.cast_or_expr())
            while self.peek()[1] == ",":
                self.next(); args.append(self.cast_or_expr())
        self.expect(")")
        return ("call", name.upper(), args)

    def cast_or_expr(self):
        # CAST(x AS numeric(…)) — zachytit AS
        node = self.expr()
        if self.peek()[1] and str(self.peek()[1]).upper() == "AS":
            self.next()
            typ = self.next()[1]  # numeric / int / …
            # volitelně (p,s)
            if self.peek()[1] == "(":
                depth = 0
                while True:
                    v = self.next()[1]
                    if v == "(":
                        depth += 1
                    elif v == ")":
                        depth -= 1
                        if depth == 0:
                            break
            return ("cast", node, (typ or "").lower())
        return node


# ── evaluator ────────────────────────────────────────────────────────────────
def _is_num(x):
    if isinstance(x, Decimal):
        return True
    if isinstance(x, str):
        try:
            Decimal(x.strip().replace(",", "."))
            return x.strip() != ""
        except (InvalidOperation, ValueError):
            return False
    return False


def _dec(x):
    if isinstance(x, Decimal):
        return x
    return Decimal(str(x).strip().replace(",", "."))


def _num_for_add(x):
    """Vrátí Decimal, pokud je operand '+' bezpečně číselný (cena), jinak None (=konkat).
    Kódy s vedoucí nulou před další číslicí ('0000000', '09120009901') a nečíselné
    řetězce/literály zůstávají řetězci → '+' je pak konkatenace (SQL varchar sémantika)."""
    if isinstance(x, Decimal):
        return x
    if isinstance(x, str):
        s = x.strip().replace(",", ".")
        if s == "" or not re.fullmatch(r"-?\d+(\.\d+)?", s):
            return None
        if re.match(r"-?0\d", s):   # '0' následované číslicí = kód (007, 0912), ne číslo
            return None
        try:
            return Decimal(s)
        except (InvalidOperation, ValueError):
            return None
    return None


def _s(x):
    if x is None:
        return ""
    if isinstance(x, Decimal):
        return format(x.normalize(), "f")
    return str(x)


def _eval(node, params):
    typ = node[0]
    if typ == "str":
        return node[1]
    if typ == "num":
        return Decimal(node[1])
    if typ == "name":
        return node[1]
    if typ == "param":
        return params.get(node[1], "")
    if typ == "cast":
        val = _eval(node[1], params)
        t = node[2]
        if t.startswith("num") or t in ("int", "bigint", "float", "money", "real", "decimal"):
            try:
                return _dec(val)
            except (InvalidOperation, ValueError):
                return Decimal(0)
        return _s(val)
    if typ == "bin":
        op = node[1]
        a = _eval(node[2], params)
        b = _eval(node[3], params)
        if op == "+":
            # SQL sémantika: sčítá jen když jsou OBA bezpečně číselné (Decimal nebo
            # číselný string bez vedoucí nuly); jinak (řetězcové literály, kódy,
            # varchar parametry) = spojení řetězců. Viz _num_for_add.
            na, nb = _num_for_add(a), _num_for_add(b)
            if na is not None and nb is not None:
                return na + nb
            return _s(a) + _s(b)
        if op == "-":
            return _dec(a) - _dec(b)
        if op == "*":
            return _dec(a) * _dec(b)
        if op == "/":
            db = _dec(b)
            return (_dec(a) / db) if db != 0 else Decimal(0)
    if typ == "null":
        return None
    if typ == "cmp":
        return _compare(_eval(node[2], params), _eval(node[3], params), node[1])
    if typ == "case":
        searched, operand, branches, elseval = node[1], node[2], node[3], node[4]
        for wp, res in branches:
            hit = _truthy(_eval(wp, params)) if searched else _eq(_eval(operand, params), _eval(wp, params))
            if hit:
                return _eval(res, params)
        return _eval(elseval, params) if elseval is not None else None
    if typ == "call":
        name, args = node[1], [_eval(a, params) for a in node[2]]
        return _fn(name, args)
    raise ValueError("neznámý uzel: %r" % (typ,))


def _compare(a, b, op):
    if _is_num(a) and _is_num(b):
        a, b = _dec(a), _dec(b)
    else:
        a, b = _s(a), _s(b)
    if op == "=":
        return a == b
    if op in ("<>", "!="):
        return a != b
    if op == ">":
        return a > b
    if op == "<":
        return a < b
    if op == ">=":
        return a >= b
    if op == "<=":
        return a <= b
    return False


def _truthy(x):
    if isinstance(x, bool):
        return x
    if x is None:
        return False
    if _is_num(x):
        return _dec(x) != 0
    return _s(x) != ""


def _eq(a, b):
    if _is_num(a) and _is_num(b):
        return _dec(a) == _dec(b)
    return _s(a) == _s(b)


def _fn(name, a):
    if name == "SUBSTRING":
        s = _s(a[0]); start = int(_dec(a[1])); ln = int(_dec(a[2]))
        return s[start - 1: start - 1 + ln] if start >= 1 else s[:0]
    if name in ("RIGHT",):
        s = _s(a[0]); n = int(_dec(a[1])); return s[-n:] if n > 0 else ""
    if name in ("LEFT",):
        s = _s(a[0]); n = int(_dec(a[1])); return s[:n]
    if name == "REPLACE":
        return _s(a[0]).replace(_s(a[1]), _s(a[2]))
    if name == "CHAR":
        return chr(int(_dec(a[0])))
    if name == "UPPER":
        return _s(a[0]).upper()
    if name == "LOWER":
        return _s(a[0]).lower()
    if name in ("LTRIM",):
        return _s(a[0]).lstrip()
    if name in ("RTRIM",):
        return _s(a[0]).rstrip()
    if name in ("TRIM",):
        return _s(a[0]).strip()
    if name in ("LEN", "LENGTH"):
        return Decimal(len(_s(a[0])))
    if name == "ISNUMERIC":
        return Decimal(1) if _is_num(a[0]) else Decimal(0)
    if name == "ROUND":
        n = int(_dec(a[1])) if len(a) > 1 else 0
        try:
            return _dec(a[0]).quantize(Decimal(1).scaleb(-n), rounding=ROUND_HALF_UP)
        except Exception:
            return _dec(a[0])
    if name == "ABS":
        return abs(_dec(a[0]))
    if name in ("FLOOR", "CEILING", "CEIL"):
        import math as _m
        d = _dec(a[0])
        return Decimal(_m.floor(d) if name == "FLOOR" else _m.ceil(d))
    if name == "ISNULL":
        return a[0] if (a[0] is not None and _s(a[0]) != "") else (a[1] if len(a) > 1 else "")
    if name == "CONCAT":
        return "".join(_s(x) for x in a)
    if name == "CAST":  # CAST bez AS (fallback)
        return a[0]
    raise ValueError("nepovolená funkce: %s" % name)


def eval_expr(expr: str, params: dict):
    """Vyhodnotí jeden vzorec. params = {'P01': hodnota, ...}. Vrací str/Decimal."""
    ast = _P(_tokenize(expr)).parse()
    return _eval(ast, params)


def transform_row(vzorce, raw_params: dict) -> dict:
    """vzorce = uspořádaný list dictů {poradi, cil_pole, vyraz}. raw_params = {'P01':…}.
    Vrací dict výstupních polí (RegCisHeo, EC_PC, EC_NC, Popis, EAN, …).
    Cíl '@Pxx' = pracovní slot (přepíše params pro další vzorce)."""
    p = dict(raw_params)
    out = {}
    for v in sorted(vzorce, key=lambda x: x.get("poradi") or 0):
        cil = (v.get("cil_pole") or "").strip()
        expr = v.get("vyraz") or ""
        if not cil or not expr:
            continue
        # SQL subquery (např. aktuální cena mědi) — evaluátor SQL neumí; hodnota
        # musí být předem naseedovaná v raw_params (viz import_by_config → @P13).
        if expr.lstrip().lstrip("(").upper().startswith("SELECT"):
            continue
        try:
            val = eval_expr(expr, p)
        except Exception as e:  # noqa: BLE001
            out.setdefault("_chyby", []).append("%s: %s" % (cil, str(e)[:80]))
            continue
        if cil.startswith("@P"):
            p[cil[1:]] = val
        else:
            out[cil] = val
    return out


def norm_kod(code: str) -> str:
    """Normalizace katalogového kódu pro párování (bez mezer, velká písmena)."""
    return re.sub(r"\s+", "", (code or "")).upper()


# ── XLS import (server-side: MCP file_read + openpyxl) ───────────────────────

def _read_share_bytes(path: str) -> bytes:
    """Přečte soubor ze sdíleného disku přes EUROSOFT MCP (base64) → bytes."""
    import base64 as _b64, json as _j, os.path as _op
    from modules.conversation.application.eurosoft_mcp_client import get_eurosoft_mcp_client
    mcp = get_eurosoft_mcp_client()
    if mcp is None:
        raise RuntimeError("EUROSOFT MCP nedostupný")
    base = _op.dirname(path)
    fn = _op.basename(path)
    raw = mcp.call_tool_sync("eurosoft_eurosoft_file_read",
                             {"user_namespace": "ro", "base_override": base,
                              "path": fn, "encoding": "base64"}, conversation_id=None)
    r = _j.loads(raw) if isinstance(raw, str) else raw
    if isinstance(r, dict) and r.get("ok") is False:
        raise RuntimeError(str(r.get("error") or r)[:200])
    b64 = (r.get("content") or r.get("data") or "") if isinstance(r, dict) else str(r)
    return _b64.b64decode(b64)


def peek_xls(path: str, n: int = 12, sheet_idx: int = 0) -> dict:
    """Náhled XLS: prvních n řádků prvního listu jako pole hodnot (pro zjištění layoutu)."""
    import io as _io
    import openpyxl as _ox
    data = _read_share_bytes(path)
    wb = _ox.load_workbook(_io.BytesIO(data), read_only=True, data_only=True)
    names = wb.sheetnames
    ws = wb[names[sheet_idx]] if sheet_idx < len(names) else wb.active
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= n:
            break
        rows.append(["" if c is None else str(c) for c in row])
    wb.close()
    maxc = max((len(r) for r in rows), default=0)
    return {"ok": True, "listy": names, "list": ws.title,
            "sloupcu": maxc, "radky": rows}


# výstupní pole vzorců (DB-Ceniky názvy) → sloupce tenant.cenik_polozka
_OUT_MAP = {
    "RegCisHeo": "kat_kod", "EC_PC": "list_price", "EC_NC": "net_price",
    "Popis": "popis", "EAN": "ean", "HmotnostKg": "hmotnost_kg",
    "MJ": "mj", "Mena": "mena", "Rabat": "rabat", "Rabat_N": "rabat",
}
_NUM_FIELDS = {"list_price", "net_price", "hmotnost_kg", "rabat"}


def _dbc_query(sql):
    """Dotaz do DB-Ceniky (EUROSOFT MSSQL) přes MCP. Vrací list dict řádků."""
    import json as _j
    from modules.conversation.application.eurosoft_mcp_client import get_eurosoft_mcp_client
    mcp = get_eurosoft_mcp_client()
    if mcp is None:
        raise RuntimeError("EUROSOFT MCP nedostupný")
    raw = mcp.call_tool_sync("eurosoft_strategie_query_raw",
                             {"sql": sql, "db_name": "DB_EC"}, conversation_id=None)
    r = _j.loads(raw) if isinstance(raw, str) else raw
    if isinstance(r, dict):
        return r.get("rows") or []
    return r or []


def migrate_supplier(dbc_id, vyrobce, nazev, pattern, mena="EUR", data_start=1, tenant_id=2):
    """Přenese vzorce + mapování z DB-Ceniky (IDCenik=dbc_id) do STRATEGIE configu.
    col_map se odvodí z EC_CenikyVzorcePar (Sloupec NN → index). Server-side, bez banneru."""
    import json as _j
    from sqlalchemy import text as _t
    from core.database_data import get_data_session

    vz = _dbc_query("SELECT Poradi, NazevCilSloupce, Vzorec FROM [DB-Ceniky].dbo.EC_CenikyVzorce "
                    "WHERE IDCenik=%d ORDER BY Poradi" % int(dbc_id))
    par = _dbc_query("SELECT TOP 1 P01,P02,P03,P04,P05,P06,P07,P08,P09,P10,P11,P12 "
                     "FROM [DB-Ceniky].dbo.EC_CenikyVzorcePar WHERE IDCenik=%d" % int(dbc_id))
    if not vz:
        return {"ok": False, "error": "chybí vzorce pro IDCenik %s" % dbc_id}
    col_map = {}
    if par:
        p = par[0]
        for i in range(1, 13):
            v = (p.get("P%02d" % i) if isinstance(p, dict) else (p[i - 1] if i - 1 < len(p) else None))
            if v and str(v).strip().startswith("Sloupec"):
                try:
                    col_map["P%02d" % i] = int(str(v).strip().replace("Sloupec", ""))
                except ValueError:
                    pass
    if not col_map:
        return {"ok": False, "error": "col_map prázdný (VzorcePar) pro %s" % dbc_id}

    s = get_data_session()
    try:
        s.execute(_t("INSERT INTO tenant.cenik_vyrobce(tenant_id,vyrobce,nazev,col_map,data_start,mena,soubor_pattern) "
                     "VALUES(:t,:v,:n,CAST(:cm AS jsonb),:ds,:m,:p) "
                     "ON CONFLICT (tenant_id,vyrobce) DO UPDATE SET col_map=EXCLUDED.col_map, nazev=EXCLUDED.nazev, "
                     "soubor_pattern=EXCLUDED.soubor_pattern, data_start=EXCLUDED.data_start, mena=EXCLUDED.mena"),
                  {"t": tenant_id, "v": vyrobce, "n": nazev, "cm": _j.dumps(col_map),
                   "ds": data_start, "m": mena, "p": pattern})
        s.execute(_t("DELETE FROM tenant.cenik_vzorec WHERE tenant_id=:t AND vyrobce=:v"),
                  {"t": tenant_id, "v": vyrobce})
        for r in vz:
            s.execute(_t("INSERT INTO tenant.cenik_vzorec(tenant_id,vyrobce,poradi,cil_pole,vyraz,je_default,aktivni) "
                         "VALUES(:t,:v,:po,:cf,:vy,true,true)"),
                      {"t": tenant_id, "v": vyrobce, "po": r.get("Poradi"),
                       "cf": r.get("NazevCilSloupce"), "vy": r.get("Vzorec")})
        s.commit()
    finally:
        s.close()
    return {"ok": True, "vyrobce": vyrobce, "col_map": col_map, "vzorcu": len(vz)}


# Plán migrace — dodavatelé s jasnou vazbou aktuální soubor ↔ DB-Ceniky import (2026).
_MIGRATE_PLAN = [
    {"dbc": 3966, "v": "HAR", "n": "Harting", "p": "harting_*"},
    {"dbc": 3963, "v": "MUR", "n": "Murr", "p": "murr_*"},
    {"dbc": 3962, "v": "SCH", "n": "Schneider", "p": "schneider_*"},
    {"dbc": 3961, "v": "LAP", "n": "LAPP", "p": "lapp_*"},
    {"dbc": 3959, "v": "PHO", "n": "Phoenix Contact", "p": "phoenixcontact_*"},
    {"dbc": 3956, "v": "EAT", "n": "Eaton", "p": "eaton_*"},
    {"dbc": 3950, "v": "WAG", "n": "WAGO", "p": "wago_*"},
    {"dbc": 3949, "v": "RIT", "n": "Rittal", "p": "rittal_*"},
]


def set_colmap_reimport(vyrobce, col_map, tenant_id=2, uid=1):
    """Ruční override mapování (když se aktuální soubor liší od DB-Ceniky verze):
    smaže staré importy dodavatele, nastaví col_map a znovu importuje. Server-side."""
    import json as _j
    from sqlalchemy import text as _t
    from core.database_data import get_data_session
    s = get_data_session()
    try:
        ids = [r[0] for r in s.execute(_t("SELECT id FROM tenant.cenik_import WHERE tenant_id=:t AND vyrobce=:v"),
                                       {"t": tenant_id, "v": vyrobce})]
        if ids:
            s.execute(_t("DELETE FROM tenant.cenik_polozka WHERE tenant_id=:t AND import_id = ANY(:i)"),
                      {"t": tenant_id, "i": ids})
            s.execute(_t("DELETE FROM tenant.cenik_import WHERE tenant_id=:t AND id = ANY(:i)"),
                      {"t": tenant_id, "i": ids})
        s.execute(_t("UPDATE tenant.cenik_vyrobce SET col_map=CAST(:cm AS jsonb) WHERE tenant_id=:t AND vyrobce=:v"),
                  {"cm": _j.dumps(col_map), "t": tenant_id, "v": vyrobce})
        s.commit()
    finally:
        s.close()
    return import_by_config(vyrobce, tenant_id=tenant_id, uid=uid)


def migrate_all(tenant_id=2, uid=1, plan=None):
    """Pro každý dodavatel: přenes vzorce z DB-Ceniky + importuj soubor. Vrací výsledky."""
    res = []
    for it in (plan or _MIGRATE_PLAN):
        try:
            m = migrate_supplier(it["dbc"], it["v"], it["n"], it["p"], tenant_id=tenant_id)
            if not m.get("ok"):
                res.append({"v": it["v"], "ok": False, "faze": "migrate", "error": m.get("error")})
                continue
            imp = import_by_config(it["v"], tenant_id=tenant_id, uid=uid)
            res.append({"v": it["v"], "col_map": m.get("col_map"), "vzorcu": m.get("vzorcu"),
                        "import_ok": imp.get("ok"), "vlozeno": imp.get("vlozeno"),
                        "soubor": imp.get("soubor"), "error": imp.get("error")})
        except Exception as e:  # noqa: BLE001
            res.append({"v": it["v"], "ok": False, "error": str(e)[:200]})
    return {"ok": True, "vysledky": res}


_INS_SQL = (
    "INSERT INTO tenant.cenik_polozka"
    "(tenant_id,import_id,radek_excel,raw,kat_kod,kat_kod_norm,popis,"
    " list_price,net_price,rabat,mj,ean,hmotnost_kg,mena) VALUES"
    "(:tenant_id,:import_id,:radek_excel,CAST(:raw AS jsonb),:kat_kod,:kat_kod_norm,"
    ":popis,:list_price,:net_price,:rabat,:mj,:ean,:hmotnost_kg,:mena)")


def _flush(s, batch):
    """Vloží dávku; při chybě fallback na per-řádek (přeskočí vadné). Odolné +
    zachytí příčinu. Vrací (vlozeno, chyb, prvni_chyba)."""
    from sqlalchemy import text as _t
    ins = _t(_INS_SQL)
    try:
        s.execute(ins, batch)
        s.commit()
        return len(batch), 0, None
    except Exception:
        s.rollback()
        ok = 0
        err = 0
        first = None
        for rec in batch:
            try:
                s.execute(ins, rec)
                s.commit()
                ok += 1
            except Exception as e:  # noqa: BLE001
                s.rollback()
                err += 1
                if first is None:
                    first = "ř.%s: %s: %s" % (rec.get("radek_excel"), type(e).__name__, str(e)[:160])
        return ok, err, first


def import_cenik(path, vyrobce, col_map, data_start=1, mena="EUR",
                 ceny_czk=False, platnost_od=None, tenant_id=2, uid=None, limit=None,
                 seed_params=None):
    """Import XLS ceníku: staging raw (JSONB) + aplikace vzorců (z tenant.cenik_vzorec
    dle vyrobce) → normalizovaná pole. col_map = {'P01': index_sloupce (1-based), ...}."""
    import io as _io
    import json as _j
    import openpyxl as _ox
    from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
    from sqlalchemy import text as _t
    from core.database_data import get_data_session

    s = get_data_session()
    try:
        vz = [dict(r) for r in s.execute(_t(
            "SELECT poradi, cil_pole, vyraz FROM tenant.cenik_vzorec "
            "WHERE tenant_id=:t AND vyrobce=:v AND aktivni ORDER BY poradi, id"),
            {"t": tenant_id, "v": vyrobce}).mappings()]
        if not vz:
            return {"ok": False, "error": "žádné vzorce pro výrobce %s" % vyrobce}

        data = _read_share_bytes(path)
        wb = _ox.load_workbook(_io.BytesIO(data), read_only=True, data_only=True)
        ws = wb.active
        allrows = list(ws.iter_rows(values_only=True))
        wb.close()

        imp = s.execute(_t(
            "INSERT INTO tenant.cenik_import(tenant_id,vyrobce,mena,ceny_czk,platnost_od,"
            "zdroj_soubor,mapovani,created_by) VALUES(:t,:v,:m,:c,:po,:zs,CAST(:mp AS jsonb),:by) "
            "RETURNING id"),
            {"t": tenant_id, "v": vyrobce, "m": mena, "c": ceny_czk, "po": platnost_od,
             "zs": path, "mp": _j.dumps({"col_map": col_map, "data_start": data_start}),
             "by": uid}).scalar()
        s.commit()  # hlavička persistuje i kdyby řádky selhaly

        cnt = 0        # zpracováno řádků
        vlozeno = 0    # reálně vloženo do DB
        db_chyb = 0    # DB chyby (přeskočené řádky)
        chyb = 0       # transform (vzorec) chyby
        first_err = None
        batch = []

        def _do_flush(_b):
            nonlocal vlozeno, db_chyb, first_err
            ok, er, fe = _flush(s, _b)
            vlozeno += ok
            db_chyb += er
            if fe and first_err is None:
                first_err = fe

        for i in range(data_start, len(allrows)):
            row = allrows[i]
            if row is None or all(c is None or str(c).strip() == "" for c in row):
                continue
            params = {str(k): ("" if v is None else str(v)) for k, v in (seed_params or {}).items()}
            for pk, ci in col_map.items():
                params[pk] = "" if (ci - 1 >= len(row) or row[ci - 1] is None) else str(row[ci - 1])
            out = transform_row(vz, params)
            if out.get("_chyby"):
                chyb += 1
            rec = {"tenant_id": tenant_id, "import_id": imp, "radek_excel": i + 1,
                   "raw": _j.dumps({("c%02d" % (k + 1)): ("" if v is None else str(v))
                                    for k, v in enumerate(row)}),
                   "kat_kod": None, "kat_kod_norm": None, "popis": None, "list_price": None,
                   "net_price": None, "rabat": None, "mj": None, "ean": None,
                   "hmotnost_kg": None, "mena": mena}
            for of, val in out.items():
                if of == "_chyby":
                    continue
                col = _OUT_MAP.get(of)
                if not col:
                    continue
                if col in _NUM_FIELDS:
                    try:
                        sv = str(val).strip().replace(",", ".")
                        rec[col] = Decimal(sv) if sv != "" else None
                    except (InvalidOperation, ValueError):
                        rec[col] = None
                else:
                    rec[col] = str(val)[:500]
            if rec["kat_kod"]:
                rec["kat_kod_norm"] = norm_kod(rec["kat_kod"])
            batch.append(rec)
            cnt += 1
            if len(batch) >= 500:
                _do_flush(batch); batch = []
            if limit and cnt >= limit:
                break
        if batch:
            _do_flush(batch)
        _pozn = None
        if db_chyb:
            _pozn = ("db_chyb=%s | %s" % (db_chyb, first_err or ""))[:500]
        s.execute(_t("UPDATE tenant.cenik_import SET pocet_polozek=:n, zpracovano=true, "
                     "poznamka=:pz, updated_at=now() WHERE id=:i"),
                  {"n": vlozeno, "pz": _pozn, "i": imp})
        s.commit()
        return {"ok": True, "import_id": imp, "vlozeno": vlozeno, "zpracovano": cnt,
                "vzorcu": len(vz), "vzorec_chyb": chyb, "db_chyb": db_chyb,
                "prvni_chyba": first_err}
    finally:
        s.close()


_CENIK_DIR = r"D:\Data\ZZ_Marti-AI RW\Ceniky"


def list_cenik_dir():
    """Názvy souborů v adresáři Ceniky (přes MCP)."""
    import json as _j
    from modules.conversation.application.eurosoft_mcp_client import get_eurosoft_mcp_client
    mcp = get_eurosoft_mcp_client()
    if mcp is None:
        return []
    raw = mcp.call_tool_sync("eurosoft_eurosoft_file_list",
                             {"user_namespace": "ro", "base_override": _CENIK_DIR, "subpath": ""},
                             conversation_id=None)
    r = _j.loads(raw) if isinstance(raw, str) else raw
    items = (r.get("items") or r.get("files") or r.get("entries") or []) if isinstance(r, dict) else (r or [])
    out = []
    for it in items:
        nm = (it.get("name") or it.get("filename") or it.get("path")) if isinstance(it, dict) else it
        if nm:
            out.append(nm)
    return out


# ── cena mědi (LAPP a další kabelové ceníky závisí na aktuálním kurzu mědi) ────
# Obdoba DB_EC.dbo.EC_CenaMedi: vzorec @P13 = TOP 1 CenaMed_EUR_100Kg ORDER BY ID DESC.
# Ve STRATEGII zrcadlíme do tenant.cenik_cena_medi a při importu seedujeme @P13.

def current_copper_price(tenant_id=2):
    """Aktuální cena mědi (EUR/100kg) = nejnovější řádek tenant.cenik_cena_medi.
    Vrací Decimal nebo None."""
    from sqlalchemy import text as _t
    from core.database_data import get_data_session
    s = get_data_session()
    try:
        r = s.execute(_t(
            "SELECT cena_eur_100kg FROM tenant.cenik_cena_medi "
            "WHERE tenant_id=:t ORDER BY datum DESC NULLS LAST, id DESC LIMIT 1"),
            {"t": tenant_id}).scalar()
        return r
    finally:
        s.close()


def sync_copper_from_ec(tenant_id=2):
    """Zrcadlí DB_EC.dbo.EC_CenaMedi → tenant.cenik_cena_medi (podle ec_id, idempotentní)."""
    from sqlalchemy import text as _t
    from core.database_data import get_data_session
    rows = _dbc_query("SELECT ID, Datum, CenaMed_EUR_100Kg, Autor, DatPorizeni "
                      "FROM DB_EC.dbo.EC_CenaMedi ORDER BY ID")
    s = get_data_session()
    ins = 0
    try:
        for r in rows:
            ec_id = r.get("ID") if isinstance(r, dict) else None
            datum = r.get("Datum") if isinstance(r, dict) else None
            cena = r.get("CenaMed_EUR_100Kg") if isinstance(r, dict) else None
            autor = r.get("Autor") if isinstance(r, dict) else None
            dpor = r.get("DatPorizeni") if isinstance(r, dict) else None
            if cena is None:
                continue
            s.execute(_t(
                "INSERT INTO tenant.cenik_cena_medi(tenant_id,ec_id,datum,cena_eur_100kg,autor,dat_porizeni,zdroj) "
                "VALUES(:t,:e,CAST(:d AS date),CAST(:c AS numeric),:a,CAST(:dp AS timestamp),'DB_EC') "
                "ON CONFLICT (tenant_id,ec_id) DO UPDATE SET datum=EXCLUDED.datum, "
                "cena_eur_100kg=EXCLUDED.cena_eur_100kg, autor=EXCLUDED.autor"),
                {"t": tenant_id, "e": ec_id, "d": str(datum)[:10] if datum else None,
                 "c": str(cena), "a": (str(autor)[:128] if autor else None),
                 "dp": str(dpor)[:19] if dpor else None})
            ins += 1
        s.commit()
        return {"ok": True, "zrcadleno": ins, "aktualni": str(current_copper_price(tenant_id))}
    finally:
        s.close()


def add_copper_price(cena, datum=None, autor="ruční", tenant_id=2):
    """Přidá nový záznam ceny mědi (EUR/100kg). ec_id=NULL (nepochází z DB_EC)."""
    from sqlalchemy import text as _t
    from core.database_data import get_data_session
    s = get_data_session()
    try:
        rid = s.execute(_t(
            "INSERT INTO tenant.cenik_cena_medi(tenant_id,ec_id,datum,cena_eur_100kg,autor,zdroj) "
            "VALUES(:t,NULL,COALESCE(CAST(:d AS date),CURRENT_DATE),CAST(:c AS numeric),:a,'STRATEGIE') "
            "RETURNING id"),
            {"t": tenant_id, "d": datum, "c": str(cena), "a": str(autor)[:128]}).scalar()
        s.commit()
        return {"ok": True, "id": rid, "cena": str(cena), "aktualni": str(current_copper_price(tenant_id))}
    finally:
        s.close()


def list_copper(tenant_id=2, limit=15):
    """Posledních N cen mědi."""
    from sqlalchemy import text as _t
    from core.database_data import get_data_session
    s = get_data_session()
    try:
        rows = [dict(r) for r in s.execute(_t(
            "SELECT id, ec_id, datum, cena_eur_100kg, autor, zdroj FROM tenant.cenik_cena_medi "
            "WHERE tenant_id=:t ORDER BY datum DESC NULLS LAST, id DESC LIMIT :n"),
            {"t": tenant_id, "n": limit}).mappings()]
        return {"ok": True, "aktualni": str(current_copper_price(tenant_id)), "radky": rows}
    finally:
        s.close()


def _copper_seed_for(vyrobce, tenant_id=2):
    """Pokud vzorce výrobce závisí na ceně mědi (SELECT ...CenaMedi), vrať {'P13': cena}."""
    from sqlalchemy import text as _t
    from core.database_data import get_data_session
    s = get_data_session()
    try:
        dep = s.execute(_t(
            "SELECT 1 FROM tenant.cenik_vzorec WHERE tenant_id=:t AND vyrobce=:v "
            "AND aktivni AND vyraz ILIKE '%%CenaMedi%%' LIMIT 1"),
            {"t": tenant_id, "v": vyrobce}).scalar()
    finally:
        s.close()
    if not dep:
        return None
    cp = current_copper_price(tenant_id)
    if cp is None:
        return None
    # cíl @P13 → param 'P13'
    return {"P13": str(cp)}


def import_by_config(vyrobce, path=None, limit=None, tenant_id=2, uid=1):
    """Generický import: načte config výrobce (col_map/data_start/mena/pattern), najde
    nejnovější soubor v adresáři Ceniky (pokud path není zadán) a spustí import_cenik."""
    import os.path as _op
    import fnmatch as _fn
    from sqlalchemy import text as _t
    from core.database_data import get_data_session
    s = get_data_session()
    try:
        cfg = s.execute(_t(
            "SELECT col_map, data_start, mena, ceny_czk, soubor_pattern, nazev "
            "FROM tenant.cenik_vyrobce WHERE tenant_id=:t AND vyrobce=:v AND aktivni"),
            {"t": tenant_id, "v": vyrobce}).mappings().first()
    finally:
        s.close()
    if not cfg:
        return {"ok": False, "error": "chybí config pro výrobce %s" % vyrobce}
    if not path:
        pat = (cfg["soubor_pattern"] or (vyrobce + "*")).lower().replace("%", "*")
        if "*" not in pat and "?" not in pat:
            pat = pat + "*"
        files = [f for f in list_cenik_dir()
                 if _fn.fnmatch(f.lower(), pat) and f.lower().endswith((".xlsx", ".xls"))]
        if not files:
            return {"ok": False, "error": "soubor nenalezen (pattern %s)" % pat}
        files.sort()
        path = _op.join(_CENIK_DIR, files[-1])
    seed = _copper_seed_for(vyrobce, tenant_id=tenant_id)
    r = import_cenik(path, vyrobce, cfg["col_map"], data_start=cfg["data_start"] or 1,
                     mena=cfg["mena"] or "EUR", ceny_czk=cfg["ceny_czk"] or False,
                     tenant_id=tenant_id, uid=uid, limit=limit, seed_params=seed)
    if isinstance(r, dict):
        r["soubor"] = _op.basename(path)
        r["nazev"] = cfg["nazev"]
    return r


# ── UI / dotazy / kalkulace ──────────────────────────────────────────────────

def prehled(tenant_id=2):
    """Přehled dodavatelů (config) + poslední import + počet položek."""
    from sqlalchemy import text as _t
    from core.database_data import get_data_session
    s = get_data_session()
    try:
        rows = s.execute(_t("""
            SELECT c.vyrobce, c.nazev, c.mena, c.aktivni,
                   (SELECT count(*) FROM tenant.cenik_vzorec v WHERE v.tenant_id=c.tenant_id AND v.vyrobce=c.vyrobce AND v.aktivni) AS vzorcu,
                   li.id AS last_import, li.pocet_polozek,
                   to_char(li.created_at,'DD.MM.YYYY HH24:MI') AS importovano
            FROM tenant.cenik_vyrobce c
            LEFT JOIN LATERAL (SELECT id, pocet_polozek, created_at FROM tenant.cenik_import i
                               WHERE i.tenant_id=c.tenant_id AND i.vyrobce=c.vyrobce
                               ORDER BY i.id DESC LIMIT 1) li ON true
            WHERE c.tenant_id=:t ORDER BY c.nazev
        """), {"t": tenant_id}).mappings().all()
        celkem = s.execute(_t("SELECT count(*) FROM tenant.cenik_polozka WHERE tenant_id=:t"),
                           {"t": tenant_id}).scalar()
        return {"ok": True, "dodavatele": [dict(r) for r in rows], "polozek_celkem": celkem}
    finally:
        s.close()


def polozky(tenant_id=2, vyrobce=None, q=None, limit=100):
    """Prohlížeč položek (poslední import daného výrobce / hledání dle kódu/popisu)."""
    from sqlalchemy import text as _t
    from core.database_data import get_data_session
    s = get_data_session()
    try:
        where = ["p.tenant_id=:t"]
        params = {"t": tenant_id, "lim": limit}
        if vyrobce:
            where.append("i.vyrobce=:v AND i.id=(SELECT max(id) FROM tenant.cenik_import WHERE tenant_id=:t AND vyrobce=:v)")
            params["v"] = vyrobce
        if q:
            where.append("(p.kat_kod_norm LIKE :q OR upper(p.popis) LIKE :q)")
            params["q"] = "%" + norm_kod(q) + "%" if not vyrobce else "%" + q.upper() + "%"
            params["q"] = "%" + q.upper().replace(" ", "") + "%"
        rows = s.execute(_t(
            "SELECT p.kat_kod, p.popis, p.list_price, p.net_price, p.rabat, p.ean, p.mena, i.vyrobce "
            "FROM tenant.cenik_polozka p JOIN tenant.cenik_import i ON i.id=p.import_id "
            "WHERE " + " AND ".join(where) + " ORDER BY p.kat_kod LIMIT :lim"), params).mappings().all()
        return {"ok": True, "polozky": [dict(r) for r in rows]}
    finally:
        s.close()


def find_price(kat_kod, tenant_id=2):
    """Kalkulace: najdi net/list cenu komponenty dle katalogového kódu (nejnovější aktivní
    import daného výrobce). Vrací i alternativy shody."""
    from sqlalchemy import text as _t
    from core.database_data import get_data_session
    kn = norm_kod(kat_kod)
    s = get_data_session()
    try:
        rows = s.execute(_t("""
            SELECT p.kat_kod, p.popis, p.net_price, p.list_price, p.rabat, p.mena, i.vyrobce,
                   i.id AS import_id, to_char(i.created_at,'DD.MM.YYYY') AS import_dne
            FROM tenant.cenik_polozka p JOIN tenant.cenik_import i ON i.id=p.import_id
            WHERE p.tenant_id=:t AND p.kat_kod_norm=:kn
              AND i.id=(SELECT max(id) FROM tenant.cenik_import WHERE tenant_id=:t AND vyrobce=i.vyrobce)
            ORDER BY i.id DESC LIMIT 5
        """), {"t": tenant_id, "kn": kn}).mappings().all()
        if not rows:
            return {"ok": True, "nalezeno": False, "kat_kod": kat_kod}
        best = dict(rows[0])
        return {"ok": True, "nalezeno": True, "cena": best, "shod": len(rows)}
    finally:
        s.close()


def dedup_imports(tenant_id=2):
    """Nechá jen NEJNOVĚJŠÍ import per výrobce, starší (+ jejich položky) smaže."""
    from sqlalchemy import text as _t
    from core.database_data import get_data_session
    s = get_data_session()
    try:
        keep = [r[0] for r in s.execute(_t(
            "SELECT max(id) FROM tenant.cenik_import WHERE tenant_id=:t GROUP BY vyrobce"),
            {"t": tenant_id})]
        if not keep:
            return {"ok": True, "smazano_importu": 0}
        old = [r[0] for r in s.execute(_t(
            "SELECT id FROM tenant.cenik_import WHERE tenant_id=:t AND id <> ALL(:keep)"),
            {"t": tenant_id, "keep": keep})]
        if old:
            s.execute(_t("DELETE FROM tenant.cenik_polozka WHERE tenant_id=:t AND import_id = ANY(:o)"),
                      {"t": tenant_id, "o": old})
            s.execute(_t("DELETE FROM tenant.cenik_import WHERE tenant_id=:t AND id = ANY(:o)"),
                      {"t": tenant_id, "o": old})
            s.commit()
        return {"ok": True, "ponechano": keep, "smazano_importu": len(old)}
    finally:
        s.close()
