"""Docházka po zakázkách — data pro vlastní stránku ve stylu standardu přehledů
(table.dokl, vzor pokladny.html). Peťa + Claude‑26, 22.7.2026.

Přehled zůstává definovaný v JEDNOM místě — data_setu `dochazka.zakazky_vse_list`
(resp. `dochazka.zakazky_budoucnost_list`). Endpoint jen vezme jeho SQL a spustí ho,
aby vlastní stránka renderovala TÁŽ data jako dřív framework grid. Jen čte (PG).

Gate = shodný s viditelností uzlu „🕒 Docházka" (11 lidí) + rodičovský bypass.
"""
from __future__ import annotations

from fastapi import APIRouter, File, Form, Request, UploadFile
from fastapi.responses import JSONResponse

doch_zak_tab_router = APIRouter(prefix="/api/v1/erp", tags=["dochazka-zak-tab"])

_DZT_ALLOWED = {1, 11, 13, 16, 17, 18, 20, 41, 107, 108, 109}
_DZT_DATASET = {
    "vse": "dochazka.zakazky_vse_list",
    "budoucnost": "dochazka.zakazky_budoucnost_list",
}


_DZT_WIDTHS_KEY = "dochazka_col_widths"


def _dzt_can_save(uid: int | None) -> bool:
    """Kdo smí měnit SDÍLENÉ výchozí šířky: rodiče (Marti/Kristý/Jirka) + Peťa (18)."""
    if not uid:
        return False
    if int(uid) in (1, 11, 20, 18):
        return True
    from sqlalchemy import text as _t
    try:
        from core.database_data import get_data_session as _g
        s = _g()
        try:
            return bool(s.execute(_t("SELECT COALESCE(is_marti_parent,false) FROM public.users WHERE id=:u"),
                                  {"u": int(uid)}).scalar())
        finally:
            s.close()
    except Exception:
        return False


def _dzt_can(uid: int | None) -> bool:
    if not uid:
        return False
    if int(uid) in _DZT_ALLOWED:
        return True
    from sqlalchemy import text as _t
    try:
        from core.database_data import get_data_session as _g
        s = _g()
        try:
            return bool(s.execute(_t("SELECT COALESCE(is_marti_parent,false) FROM public.users WHERE id=:u"),
                                  {"u": int(uid)}).scalar())
        finally:
            s.close()
    except Exception:
        return False


_DOW_CZ = ['Po', 'Ut', 'St', 'Ct', 'Pa', 'So', 'Ne']


def _dzt_ec_office_hist(s) -> list[dict]:
    """Kancelář (lidé BEZ výrobní práce) leden–květen 2026 přímo z Centrály
    (EC_Dochazka, read-only). Výroba je už v PG (vyroba_work), tu z EC vynecháme.
    Vrací řádky ve stejném tvaru jako přehled, kind='C' (jen ke čtení)."""
    from sqlalchemy import text as _t
    prod = [int(r[0]) for r in s.execute(_t(
        "SELECT DISTINCT cislo_zam FROM tenant.vyroba_work "
        "WHERE tenant_id=2 AND source_system IN ('app','centrala1') "
        "AND datum>='2026-01-01' AND cislo_zam ~ '^[0-9]+$'")).all()]
    jmena = {}
    for cz, nm in s.execute(_t(
            "SELECT cislo_zam, full_name FROM tenant.att_employee "
            "WHERE tenant_id=2 AND cislo_zam ~ '^[0-9]+$'")).all():
        try:
            jmena[int(cz)] = nm
        except Exception:
            pass
    ciny = {}
    for ec, nm in s.execute(_t(
            "SELECT ec_cislo, name FROM tenant.vyroba_cinnost WHERE ec_cislo IS NOT NULL")).all():
        try:
            ciny[int(ec)] = nm
        except Exception:
            pass
    prod_clause = ""
    if prod:
        prod_clause = " AND CisloZam NOT IN (" + ",".join(str(p) for p in prod) + ")"
    q = ("SELECT ID, DatumPripadu, CisloZam, DruhCinnosti, CisloZakazky, "
         "CasZacatek, CasKonec, CasCelkemVcRezii, ISNULL(ZamPoznamka,'') AS pozn "
         "FROM dbo.EC_Dochazka "
         "WHERE DatumPripadu>='2026-01-01' AND DatumPripadu<'2026-06-01'" + prod_clause +
         " ORDER BY DatumPripadu DESC, CasZacatek")
    out = []
    from modules.eurosoft_mcp.sql_client import get_cursor
    with get_cursor("DB_EC") as cur:
        cur.execute(q)
        cols = [d[0] for d in cur.description]
        for rec in cur.fetchall():
            r = {cols[i]: rec[i] for i in range(len(cols))}
            d = r.get("DatumPripadu")
            if d is None:
                continue
            try:
                cz = int(r.get("CisloZam"))
            except Exception:
                cz = None
            od = r.get("CasZacatek")
            kon = r.get("CasKonec")
            hv = r.get("CasCelkemVcRezii")
            hod = round(float(hv), 2) if hv is not None else None
            od_t = od.strftime("%H:%M") if od else None
            kon_t = kon.strftime("%H:%M") if kon else None
            den_iso = d.strftime("%Y-%m-%d")
            jm = jmena.get(cz) or ("Zam " + str(cz))
            dc = r.get("DruhCinnosti")
            cin = ciny.get(int(dc)) if dc is not None else None
            zak = r.get("CisloZakazky") or "Rezie"
            pozn = r.get("pozn") or ""
            out.append({
                "PraceAktivni": "", "CisloZakazky": zak, "JmenoPrijmeni": jm,
                "CisloZam": cz, "DruhCinnosti": dc, "CinnostText": cin,
                "DenVTydnu": _DOW_CZ[d.weekday()],
                "CasZacatek": d.strftime("%d.%m.%Y") + ((" " + od_t) if od_t else ""),
                "CasKonec": (kon.strftime("%d.%m.%Y %H:%M") if kon else None),
                "CasCelkem": hod, "Odkud": "z Centrály", "Smlouva": None,
                "Poznamka": pozn, "DatumPripadu": d.strftime("%d.%m.%Y"),
                "Rok": d.year, "Mesic": d.month,
                "_kind": "C", "_id": int(r.get("ID")) if r.get("ID") is not None else None,
                "_zak": zak, "_cin_id": None,
                "_od_d": (od.strftime("%Y-%m-%d") if od else den_iso), "_od_t": od_t,
                "_kon_d": (kon.strftime("%Y-%m-%d") if kon else None), "_kon_t": kon_t,
                "_hod": hod, "_pozn": pozn, "_src": "z Centrály", "_cz": cz, "_jm": jm,
            })
    return out


@doch_zak_tab_router.get("/app/dochazka-zak-tab/data")
def dochazka_zak_tab_data(req: Request) -> JSONResponse:
    """Řádky přehledu docházky po zakázkách (obdobi=vse|budoucnost)."""
    from modules.erp.api.router import _uid_from_token_or_cookie
    uid = _uid_from_token_or_cookie(req)
    if not uid:
        return JSONResponse({"ok": False, "error": "unauthorized"}, status_code=401)
    if not _dzt_can(uid):
        return JSONResponse({"ok": False, "error": "forbidden"}, status_code=403)
    obdobi = (req.query_params.get("obdobi") or "vse").strip().lower()
    # 'all' (Vše) = stejná data jako 'vse', jen bez omezení na poslední 2 měsíce
    base = "budoucnost" if obdobi == "budoucnost" else "vse"
    ds_code = _DZT_DATASET[base]
    from sqlalchemy import text as _t
    from modules.strategie_pg.application import service as _pg
    cm = _pg.get_session()
    s = cm.__enter__()
    try:
        sql = s.execute(_t("SELECT sql_text FROM fw.data_set WHERE code=:c"), {"c": ds_code}).scalar()
        if not sql:
            return JSONResponse({"ok": False, "error": "dataset_missing"}, status_code=500)
        if obdobi == "all":
            # odstraň spodní mez „poslední 2 měsíce" → zůstane celé (d <= CURRENT_DATE)
            sql = sql.replace(
                "AND d >= (date_trunc('month', CURRENT_DATE) - INTERVAL '1 month')::date", "")
        rows = s.execute(_t(sql)).mappings().all()
        from decimal import Decimal as _D
        import datetime as _dt

        def _conv(v):
            if isinstance(v, _D):
                return float(v)
            if isinstance(v, (_dt.date, _dt.datetime)):
                return v.isoformat()
            return v
        out = [{k: _conv(v) for k, v in dict(r).items()} for r in rows]
        return JSONResponse({"ok": True, "obdobi": obdobi, "pocet": len(out), "rows": out})
    except Exception as exc:  # noqa: BLE001
        return JSONResponse({"ok": False, "error": str(exc)[:200]}, status_code=500)
    finally:
        try:
            cm.__exit__(None, None, None)
        except Exception:
            pass


@doch_zak_tab_router.get("/app/dochazka-zak-tab/widths")
def dochazka_zak_tab_widths_get(req: Request) -> JSONResponse:
    """Šířky sloupců: `base` = sdílené výchozí pro všechny; `me` = osobní uložené
    tažení tohoto uživatele (má přednost). Osobní se ukládají do DB (jako dřív
    framework grid), aby je Claude viděl a mohl je povýšit na výchozí."""
    from modules.erp.api.router import _uid_from_token_or_cookie
    uid = _uid_from_token_or_cookie(req)
    if not uid or not _dzt_can(uid):
        return JSONResponse({"ok": True, "base": {}, "me": {}})
    from sqlalchemy import text as _t
    try:
        from core.database_data import get_data_session as _g
        s = _g()
        try:
            base = s.execute(_t("SELECT hodnota FROM tenant.att_ui_pref WHERE kod=:k"),
                             {"k": _DZT_WIDTHS_KEY}).scalar()
            me = s.execute(_t("SELECT hodnota FROM tenant.att_ui_pref WHERE kod=:k"),
                           {"k": _DZT_WIDTHS_KEY + "_u" + str(int(uid))}).scalar()
        finally:
            s.close()
        return JSONResponse({"ok": True, "base": base or {}, "me": me or {}})
    except Exception:
        return JSONResponse({"ok": True, "base": {}, "me": {}})


@doch_zak_tab_router.get("/app/dochazka-zak-tab/cinnosti")
def dochazka_zak_tab_cinnosti(req: Request) -> JSONResponse:
    """Seznam činností pro roletku ve formuláři úprav (id, ec_cislo, název)."""
    from modules.erp.api.router import _uid_from_token_or_cookie
    uid = _uid_from_token_or_cookie(req)
    if not uid or not _dzt_can(uid):
        return JSONResponse({"ok": False, "error": "forbidden"}, status_code=403)
    from sqlalchemy import text as _t
    try:
        from core.database_data import get_data_session as _g
        s = _g()
        try:
            rows = s.execute(_t(
                "SELECT id, ec_cislo, name FROM tenant.vyroba_cinnost "
                "WHERE COALESCE(name,'')<>'' AND COALESCE(active,true) "
                "ORDER BY name")).mappings().all()
        finally:
            s.close()
        out = [{"id": int(r["id"]), "ec": r["ec_cislo"], "name": r["name"]} for r in rows]
        return JSONResponse({"ok": True, "cinnosti": out})
    except Exception as exc:  # noqa: BLE001
        return JSONResponse({"ok": False, "error": str(exc)[:200]}, status_code=500)


@doch_zak_tab_router.get("/app/dochazka-zak-tab/zakazky")
def dochazka_zak_tab_zakazky(req: Request) -> JSONResponse:
    """Píchatelné zakázky pro našeptávač Zakázky ve formuláři (cislo, nazev, typ).
    Stejný zdroj jako Opravy docházky (tenant.zakazka, pichatelna=true), REZIE první."""
    from modules.erp.api.router import _uid_from_token_or_cookie
    uid = _uid_from_token_or_cookie(req)
    if not uid or not _dzt_can(uid):
        return JSONResponse({"ok": False, "error": "forbidden"}, status_code=403)
    from sqlalchemy import text as _t
    try:
        from core.database_data import get_data_session as _g
        s = _g()
        try:
            rows = s.execute(_t(
                "SELECT cislo, COALESCE(nazev,'') AS nazev, COALESCE(typ,'') AS typ "
                "FROM tenant.zakazka WHERE tenant_id=2 AND pichatelna=true "
                "ORDER BY (typ='REZIE') DESC, cislo")).mappings().all()
        finally:
            s.close()
        out = [{"cislo": r["cislo"], "nazev": r["nazev"], "typ": r["typ"]} for r in rows]
        return JSONResponse({"ok": True, "zakazky": out})
    except Exception as exc:  # noqa: BLE001
        return JSONResponse({"ok": False, "error": str(exc)[:200]}, status_code=500)


# ── Kontrolní přehledy (soudek „Kontrolní přehledy" pod Docházkou) ──────────
# report=fpd     → Hlídání FPD (HPP): odpracováno vs má-být (úvazek×prac.dny), chybí>0.
# report=prekryv → Překrytá docházka: časové překryvy záznamů téhož člověka.
# Vzor přehledů 1088 / 1084 z Centrály. Odpracováno = vyroba_work (práce+režie) +
# placená absence (dovolená/nemoc…). POZN.: výpočet hodin sjednotí Jirka (viz jeho
# nález o rozdílných výpočtech) — pak se sem napojíme; zatím vlastní výpočet.
_KONTROLA_FPD_SQL = """
WITH per AS (
  SELECT (CASE WHEN extract(day FROM current_date) < 13
            THEN date_trunc('month', current_date) - interval '1 month'
            ELSE date_trunc('month', current_date) END)::date AS m_od),
p AS (SELECT m_od, LEAST(current_date, (m_od + interval '1 month')::date) AS ke_dni FROM per),
pracdny AS (SELECT count(*) AS pocet FROM tenant.att_calendar_day cd, p
  WHERE cd.tenant_id=2 AND cd.is_workday AND COALESCE(cd.is_holiday,false)=false
    AND cd.day >= p.m_od AND cd.day < p.ke_dni),
base AS (
  SELECT em.cislo_zam, em.full_name,
    (COALESCE((SELECT sum(w.hodiny) FROM tenant.vyroba_work w, p
        WHERE w.tenant_id=2 AND w.user_id=em.user_id AND w.is_active
          AND w.datum >= p.m_od AND w.datum < p.ke_dni),0)
     + COALESCE((SELECT sum(e.hours) FROM tenant.att_entry e
          JOIN tenant.att_entry_type et ON et.id=e.entry_type_id, p
        WHERE e.tenant_id=2 AND e.employee_id=em.id AND et.category='absence' AND et.is_paid
          AND COALESCE(e.status,'')<>'superseded'
          AND e.entry_date >= p.m_od AND e.entry_date < p.ke_dni),0)) AS odpracovano,
    (SELECT pocet FROM pracdny) AS prac_dny,
    (g.uvazek_tyden_h / NULLIF(COALESCE(wm.dny_v_tydnu,5),0)) AS hod_den
  FROM tenant.att_employee em
  JOIN tenant.engagement g ON g.employee_id=em.id AND g.tenant_id=2 AND g.is_current=true
  LEFT JOIN tenant.work_mode wm ON wm.id=g.work_mode_id
  WHERE em.tenant_id=2 AND UPPER(COALESCE(g.engagement_type,''))='HPP'
    AND g.uvazek_tyden_h IS NOT NULL
    AND em.cislo_zam NOT IN ('21','41','15','2','47','361'))
SELECT full_name AS "JmenoPrijmeni", cislo_zam AS "CisloZam",
       round(odpracovano::numeric,2) AS "Odpracovano",
       prac_dny AS "PracDny", round(hod_den::numeric,2) AS "HodDen",
       round((hod_den*prac_dny)::numeric,2) AS "MaByt",
       round((hod_den*prac_dny - odpracovano)::numeric,2) AS "Chybi",
       extract(month FROM (SELECT m_od FROM per))::int AS "Mesic",
       to_char((SELECT ke_dni FROM p) - 1,'DD.MM.YYYY') AS "KeDni"
FROM base
WHERE (hod_den*prac_dny - odpracovano) > 0.01
ORDER BY (hod_den*prac_dny - odpracovano) DESC
"""

_KONTROLA_PREKRYV_SQL = """
WITH prace AS (
  SELECT w.id, w.cislo_zam, COALESCE(em.full_name,'Zam '||w.cislo_zam) AS jmeno, w.od, w.konec
  FROM tenant.vyroba_work w
  LEFT JOIN tenant.att_employee em ON em.tenant_id=2 AND em.user_id=w.user_id
  WHERE w.tenant_id=2 AND w.is_active AND w.od IS NOT NULL AND w.konec IS NOT NULL
    AND w.datum >= '2026-01-01'),
prc AS (
  SELECT DISTINCT a.id, 'práce × práce' AS typ, a.cislo_zam, a.jmeno, a.od, a.konec
  FROM prace a JOIN prace b
    ON b.id<>a.id AND b.cislo_zam=a.cislo_zam AND b.od >= a.od AND b.od < a.konec),
pauzy AS (
  SELECT e.id, e.employee_id, em.cislo_zam, COALESCE(em.full_name,'Zam '||em.cislo_zam) AS jmeno,
         e.started_at AS od, e.ended_at AS konec
  FROM tenant.att_entry e JOIN tenant.att_entry_type et ON et.id=e.entry_type_id
  JOIN tenant.att_employee em ON em.id=e.employee_id AND em.tenant_id=2
  WHERE e.tenant_id=2 AND et.code='break' AND e.started_at IS NOT NULL AND e.ended_at IS NOT NULL
    AND COALESCE(e.status,'')<>'superseded' AND e.entry_date >= '2026-01-01'),
prp AS (
  SELECT DISTINCT a.id, 'přestávka × přestávka' AS typ, a.cislo_zam, a.jmeno, a.od, a.konec
  FROM pauzy a JOIN pauzy b
    ON b.id<>a.id AND b.employee_id=a.employee_id AND b.od >= a.od AND b.od < a.konec)
SELECT typ AS "Typ", cislo_zam AS "CisloZam", jmeno AS "JmenoPrijmeni",
       to_char(od,'DD.MM.YYYY HH24:MI') AS "Zacatek",
       to_char(konec,'DD.MM.YYYY HH24:MI') AS "Konec"
FROM (SELECT * FROM prc UNION ALL SELECT * FROM prp) x
ORDER BY "JmenoPrijmeni", od DESC
"""

# report=rozpad → Docházka × rozpad: sedí součet hodin za den v tenant.att_entry
# (co vidí Opravy docházky) s tenant.vyroba_work (co vidí Docházka new)? Zadala Týnka
# 31.7.2026 („ať máme kontrolu, a doufám, že se to dít nebude"). Prázdný přehled = OK.
# Období: rolující 2 měsíce, ale NIKDY před 1.7.2026 — leden–květen jsou z Centrály
# jiným postupem a červen je přechodový, tam by srovnání jen šumělo.
# Dnešek vynechán schválně (rozdělané směny nejsou chyba).
_KONTROLA_ROZPAD_SQL = """
WITH obd AS (
  SELECT GREATEST((date_trunc('month', current_date) - interval '1 month')::date,
                  DATE '2026-07-01') AS od,
         (current_date - 1) AS do),
vw AS (
  SELECT em.user_id, w.datum, sum(COALESCE(w.hodiny,0)) AS h
  FROM tenant.vyroba_work w
  JOIN tenant.att_employee em ON em.tenant_id=2 AND em.user_id=w.user_id, obd
  WHERE w.tenant_id=2 AND w.is_active AND w.datum BETWEEN obd.od AND obd.do
  GROUP BY em.user_id, w.datum),
ae AS (
  SELECT em.user_id, e.entry_date AS datum, sum(COALESCE(e.hours,0)) AS h,
         bool_or(COALESCE(e.note,'') ILIKE '%auto-odhlášení%'
                 AND to_char(e.ended_at,'HH24:MI')='23:59') AS auto_konec
  FROM tenant.att_entry e
  JOIN tenant.att_entry_type et ON et.id=e.entry_type_id AND et.category='presence'
  JOIN tenant.att_employee em ON em.id=e.employee_id AND em.tenant_id=2, obd
  WHERE e.tenant_id=2 AND et.code IN ('work','overhead','homeoffice')
    AND COALESCE(e.status,'') <> 'superseded' AND COALESCE(e.source,'') <> 'plan_ec'
    AND e.entry_date BETWEEN obd.od AND obd.do
  GROUP BY em.user_id, e.entry_date),
spolu AS (
  SELECT COALESCE(vw.user_id, ae.user_id) AS user_id,
         COALESCE(vw.datum, ae.datum) AS datum,
         round(COALESCE(vw.h,0)::numeric,2) AS h_vw,
         round(COALESCE(ae.h,0)::numeric,2) AS h_ae,
         COALESCE(ae.auto_konec,false) AS auto_konec
  FROM vw FULL OUTER JOIN ae ON ae.user_id=vw.user_id AND ae.datum=vw.datum)
SELECT (SELECT COALESCE(em.full_name,'Zam '||em.cislo_zam) FROM tenant.att_employee em
         WHERE em.tenant_id=2 AND em.user_id=spolu.user_id LIMIT 1)     AS "JmenoPrijmeni",
       (SELECT em.cislo_zam FROM tenant.att_employee em
         WHERE em.tenant_id=2 AND em.user_id=spolu.user_id LIMIT 1)     AS "CisloZam",
       to_char(datum,'DD.MM.YYYY')                                      AS "Den",
       h_ae                                                             AS "Dochazka",
       h_vw                                                             AS "Rozpad",
       round(h_vw - h_ae, 2)                                            AS "Rozdil",
       CASE WHEN auto_konec THEN 'zapomenutý odchod — konec dopsal automat'
            WHEN h_vw = 0   THEN 'rozpad na zakázky chybí úplně'
            WHEN h_ae = 0   THEN 'rozpad je, ale docházka k němu chybí'
            WHEN h_vw > h_ae THEN 'rozpad má hodiny navíc'
            ELSE 'rozpad má míň než docházka' END                       AS "Co"
FROM spolu
WHERE abs(h_vw - h_ae) > 0.1
ORDER BY abs(h_vw - h_ae) DESC, datum DESC
"""

_KONTROLA_COLS = {
    "rozpad": [
        {"k": "JmenoPrijmeni", "h": "Příjmení Jméno", "w": 180},
        {"k": "CisloZam", "h": "Číslo", "w": 60, "r": 1, "num": 1},
        {"k": "Den", "h": "Den", "w": 100},
        {"k": "Dochazka", "h": "Docházka (Opravy)", "w": 120, "r": 1, "num": 1},
        {"k": "Rozpad", "h": "Rozpad (Docházka new)", "w": 140, "r": 1, "num": 1},
        {"k": "Rozdil", "h": "Rozdíl", "w": 90, "r": 1, "num": 1},
        {"k": "Co", "h": "Co se stalo", "w": 260},
    ],
    "fpd": [
        {"k": "JmenoPrijmeni", "h": "Příjmení Jméno", "w": 170},
        {"k": "CisloZam", "h": "Číslo", "w": 60, "r": 1, "num": 1},
        {"k": "Odpracovano", "h": "Odpracováno", "w": 100, "r": 1, "num": 1},
        {"k": "PracDny", "h": "Prac. dny", "w": 70, "r": 1, "num": 1},
        {"k": "HodDen", "h": "Hod/den", "w": 70, "r": 1, "num": 1},
        {"k": "MaByt", "h": "Má být", "w": 90, "r": 1, "num": 1},
        {"k": "Chybi", "h": "Chybí odpracovat", "w": 120, "r": 1, "num": 1},
        {"k": "Mesic", "h": "Měsíc", "w": 56, "r": 1, "num": 1},
        {"k": "KeDni", "h": "Ke dni", "w": 92},
    ],
    "prekryv": [
        {"k": "Typ", "h": "Typ překryvu", "w": 150},
        {"k": "CisloZam", "h": "Číslo", "w": 60, "r": 1, "num": 1},
        {"k": "JmenoPrijmeni", "h": "Příjmení Jméno", "w": 180},
        {"k": "Zacatek", "h": "Začátek", "w": 130},
        {"k": "Konec", "h": "Konec", "w": 130},
    ],
}
_KONTROLA_TITLE = {"fpd": "Hlídání FPD (HPP)", "prekryv": "Překrytá docházka",
                   "rozpad": "Docházka × rozpad (nesedící součty)"}


@doch_zak_tab_router.get("/app/dochazka-kontrola/data")
def dochazka_kontrola_data(req: Request) -> JSONResponse:
    """Kontrolní přehledy: report=fpd | prekryv. Vrací columns + rows (jen čte, PG)."""
    from modules.erp.api.router import _uid_from_token_or_cookie
    uid = _uid_from_token_or_cookie(req)
    if not uid:
        return JSONResponse({"ok": False, "error": "unauthorized"}, status_code=401)
    if not _dzt_can(uid):
        return JSONResponse({"ok": False, "error": "forbidden"}, status_code=403)
    report = (req.query_params.get("report") or "fpd").strip().lower()
    if report not in ("fpd", "prekryv", "rozpad"):
        return JSONResponse({"ok": False, "error": "neznámý report"}, status_code=400)
    sql = {"fpd": _KONTROLA_FPD_SQL, "prekryv": _KONTROLA_PREKRYV_SQL,
           "rozpad": _KONTROLA_ROZPAD_SQL}[report]
    from sqlalchemy import text as _t
    try:
        from core.database_data import get_data_session as _g
        s = _g()
        try:
            rows = s.execute(_t(sql)).mappings().all()
        finally:
            s.close()
        from decimal import Decimal as _D
        import datetime as _dt

        def _conv(v):
            if isinstance(v, _D):
                return float(v)
            if isinstance(v, (_dt.date, _dt.datetime)):
                return v.isoformat()
            return v
        out = [{k: _conv(v) for k, v in dict(r).items()} for r in rows]
        return JSONResponse({"ok": True, "report": report,
                             "title": _KONTROLA_TITLE[report],
                             "columns": _KONTROLA_COLS[report],
                             "pocet": len(out), "rows": out})
    except Exception as exc:  # noqa: BLE001
        return JSONResponse({"ok": False, "error": str(exc)[:200]}, status_code=500)


@doch_zak_tab_router.get("/app/dochazka-zak-tab/zamestnanci")
def dochazka_zak_tab_zamestnanci(req: Request) -> JSONResponse:
    """Seznam pracovníků pro výběr pole Pracovník ve formuláři nového záznamu.
    Jen ti, které umí save-new dohledat (att_employee, user_id NOT NULL,
    číselné cislo_zam). Vrací cislo_zam + jmeno, řazeno dle jména."""
    from modules.erp.api.router import _uid_from_token_or_cookie
    uid = _uid_from_token_or_cookie(req)
    if not uid or not _dzt_can(uid):
        return JSONResponse({"ok": False, "error": "forbidden"}, status_code=403)
    from sqlalchemy import text as _t
    try:
        from core.database_data import get_data_session as _g
        s = _g()
        try:
            # ŘAZENÍ PODLE PŘÍJMENÍ (Peťa 31.7.2026). `full_name` na to nestačí —
            # je zapsané nejednotně: většinou „Jméno Příjmení" (Dušan Havlát), ale
            # i obráceně (Vlková Klára, Senft Ondřej) a s dovětky (Šafránková ml).
            # Spolehlivé příjmení je na účtu (public.users.last_name), proto join.
            # Fallback na full_name, kdyby účet příjmení neměl.
            rows = s.execute(_t(
                "SELECT em.cislo_zam, COALESCE(em.full_name,'') AS jmeno, "
                "       COALESCE(NULLIF(TRIM(u.last_name),''), em.full_name, '') AS prijmeni, "
                "       COALESCE(NULLIF(TRIM(u.first_name),''), '') AS krestni "
                "FROM tenant.att_employee em "
                "LEFT JOIN public.users u ON u.id = em.user_id "
                "WHERE em.tenant_id=2 AND em.user_id IS NOT NULL "
                "  AND em.cislo_zam ~ '^[0-9]+$' "
                "ORDER BY lower(COALESCE(NULLIF(TRIM(u.last_name),''), em.full_name, '')), "
                "         lower(COALESCE(u.first_name,''))")).mappings().all()
        finally:
            s.close()
        out = [{"cislo": r["cislo_zam"], "jmeno": r["jmeno"],
                "prijmeni": r["prijmeni"], "krestni": r["krestni"],
                # co se ukazuje v seznamu: „Příjmení Jméno" (jako v Centrále)
                "popis": (str(r["prijmeni"] or "").strip()
                          + ((" " + str(r["krestni"]).strip()) if r["krestni"] else "")).strip()
                         or (r["jmeno"] or "")} for r in rows]
        return JSONResponse({"ok": True, "zamestnanci": out})
    except Exception as exc:  # noqa: BLE001
        return JSONResponse({"ok": False, "error": str(exc)[:200]}, status_code=500)


def _dzt_parse_ts(d: str | None, t: str | None) -> str | None:
    """'YYYY-MM-DD' + 'HH:MM' → 'YYYY-MM-DD HH:MM' (None když chybí datum)."""
    d = (d or "").strip()
    if not d:
        return None
    t = (t or "").strip() or "00:00"
    return d + " " + t


@doch_zak_tab_router.post("/app/dochazka-zak-tab/save")
async def dochazka_zak_tab_save(req: Request) -> JSONResponse:
    """Uloží opravu řádku práce (tenant.vyroba_work). Edituje se jen kind='W'
    (práce na zakázce, app i Centrála). Absence (kind='A') se tu neupravují.
    Ukládají se: zakazka_ref, cinnost_id, od, konec, hodiny, poznamka.
    Pauza/Vedoucí poznámka/zaškrtávátka zatím nemají v DB místo — dodělá se."""
    from modules.erp.api.router import _uid_from_token_or_cookie
    uid = _uid_from_token_or_cookie(req)
    if not uid:
        return JSONResponse({"ok": False, "error": "unauthorized"}, status_code=401)
    if not _dzt_can(uid):
        return JSONResponse({"ok": False, "error": "forbidden"}, status_code=403)
    try:
        body = await req.json()
    except Exception:
        body = {}
    kind = str((body or {}).get("kind") or "").upper()
    try:
        rid = int((body or {}).get("id"))
    except Exception:
        return JSONResponse({"ok": False, "error": "chybí id záznamu"}, status_code=400)
    if kind != "W":
        return JSONResponse({"ok": False, "error": "Tady lze upravit jen práci na zakázce. "
                             "Absence (dovolená/nemoc/lékař) se opravují v modulu Opravy docházky."},
                            status_code=400)
    zak = (str((body or {}).get("zakazka") or "").strip()) or None
    cin_raw = (body or {}).get("cinnost_id")
    try:
        cin_id = int(cin_raw) if cin_raw not in (None, "", "null") else None
    except Exception:
        cin_id = None
    od = _dzt_parse_ts((body or {}).get("od_d"), (body or {}).get("od_t"))
    if not od:
        return JSONResponse({"ok": False, "error": "Vyplň Začátek (datum a čas)."}, status_code=400)
    kon = _dzt_parse_ts((body or {}).get("kon_d"), (body or {}).get("kon_t"))
    pozn = (body or {}).get("poznamka")
    if pozn is not None:
        pozn = str(pozn)[:2000]
    # hodiny: klient posílá spočtený Čas; jinak dopočítáme z (konec-od). Nezáporné.
    hod = (body or {}).get("hodiny")
    try:
        hod = float(hod) if hod not in (None, "") else None
    except Exception:
        hod = None
    from sqlalchemy import text as _t
    from modules.strategie_pg.application import service as _pg
    cm = _pg.get_session()
    s = cm.__enter__()
    try:
        # kontrola: řádek existuje a je to práce v našem tenantu
        cur = s.execute(_t("SELECT source_system, od, konec, hodiny FROM tenant.vyroba_work "
                           "WHERE id=:id AND tenant_id=2"), {"id": rid}).mappings().first()
        if not cur:
            return JSONResponse({"ok": False, "error": "Záznam nenalezen."}, status_code=404)
        # dopočet hodin, když je nezadal klient a máme konec
        if hod is None and kon is not None:
            r2 = s.execute(_t("SELECT EXTRACT(EPOCH FROM ((:kon)::timestamptz-(:od)::timestamptz))/3600.0"),
                           {"od": od, "kon": kon}).scalar()
            try:
                hod = round(float(r2), 2)
            except Exception:
                hod = None
        if hod is not None and hod < 0:
            return JSONResponse({"ok": False, "error": "Konec je před začátkem — zkontroluj časy."},
                                status_code=400)
        s.execute(_t(
            "UPDATE tenant.vyroba_work SET "
            " zakazka_ref=:zak, cinnost_id=:cin, od=(:od)::timestamptz, "
            " konec=CASE WHEN :kon IS NULL THEN NULL ELSE (:kon)::timestamptz END, "
            " hodiny=:hod, poznamka=:pozn, updated_at=now() "
            "WHERE id=:id AND tenant_id=2"),
            {"zak": zak, "cin": cin_id, "od": od, "kon": kon, "hod": hod,
             "pozn": pozn, "id": rid})
        s.commit()
        return JSONResponse({"ok": True, "id": rid, "hodiny": hod})
    except Exception as exc:  # noqa: BLE001
        try:
            s.rollback()
        except Exception:
            pass
        return JSONResponse({"ok": False, "error": str(exc)[:200]}, status_code=500)
    finally:
        try:
            cm.__exit__(None, None, None)
        except Exception:
            pass


@doch_zak_tab_router.post("/app/dochazka-zak-tab/save-new")
async def dochazka_zak_tab_save_new(req: Request) -> JSONResponse:
    """Založí NOVÝ řádek práce (tenant.vyroba_work) pro osobu dle cislo_zam.
    source_system='app'. Vrací id nového řádku."""
    from modules.erp.api.router import _uid_from_token_or_cookie
    uid = _uid_from_token_or_cookie(req)
    if not uid:
        return JSONResponse({"ok": False, "error": "unauthorized"}, status_code=401)
    if not _dzt_can(uid):
        return JSONResponse({"ok": False, "error": "forbidden"}, status_code=403)
    try:
        body = await req.json()
    except Exception:
        body = {}
    cz = str((body or {}).get("cislo_zam") or "").strip()
    if not cz:
        return JSONResponse({"ok": False, "error": "Chybí pracovník."}, status_code=400)
    zak = (str((body or {}).get("zakazka") or "").strip()) or None
    cin_raw = (body or {}).get("cinnost_id")
    try:
        cin_id = int(cin_raw) if cin_raw not in (None, "", "null") else None
    except Exception:
        cin_id = None
    od = _dzt_parse_ts((body or {}).get("od_d"), (body or {}).get("od_t"))
    if not od:
        return JSONResponse({"ok": False, "error": "Vyplň Začátek (datum a čas)."}, status_code=400)
    kon = _dzt_parse_ts((body or {}).get("kon_d"), (body or {}).get("kon_t"))
    pozn = (body or {}).get("poznamka")
    if pozn is not None:
        pozn = str(pozn)[:2000]
    hod = (body or {}).get("hodiny")
    try:
        hod = float(hod) if hod not in (None, "") else None
    except Exception:
        hod = None
    from sqlalchemy import text as _t
    from modules.strategie_pg.application import service as _pg
    cm = _pg.get_session()
    s = cm.__enter__()
    try:
        emp_uid = s.execute(_t("SELECT user_id FROM tenant.att_employee "
                               "WHERE tenant_id=2 AND cislo_zam=:cz AND user_id IS NOT NULL LIMIT 1"),
                            {"cz": cz}).scalar()
        if not emp_uid:
            return JSONResponse({"ok": False, "error": "Pracovníka (č. " + cz + ") se nepodařilo najít."},
                                status_code=400)
        if hod is None and kon is not None:
            r2 = s.execute(_t("SELECT EXTRACT(EPOCH FROM ((:kon)::timestamptz-(:od)::timestamptz))/3600.0"),
                           {"od": od, "kon": kon}).scalar()
            try:
                hod = round(float(r2), 2)
            except Exception:
                hod = None
        if hod is not None and hod < 0:
            return JSONResponse({"ok": False, "error": "Konec je před začátkem — zkontroluj časy."},
                                status_code=400)
        new_id = s.execute(_t(
            "INSERT INTO tenant.vyroba_work "
            " (tenant_id, user_id, cislo_zam, datum, od, konec, zakazka_ref, cinnost_id, "
            "  hodiny, poznamka, source_system, created_by, created_at, updated_at) "
            "VALUES (2, :uid, :cz, (:od)::timestamptz::date, (:od)::timestamptz, "
            "  CASE WHEN :kon IS NULL THEN NULL ELSE (:kon)::timestamptz END, "
            "  :zak, :cin, :hod, :pozn, 'app', :creator, now(), now()) RETURNING id"),
            {"uid": emp_uid, "cz": cz, "od": od, "kon": kon, "zak": zak, "cin": cin_id,
             "hod": hod, "pozn": pozn, "creator": uid}).scalar()
        s.commit()
        return JSONResponse({"ok": True, "id": new_id, "hodiny": hod})
    except Exception as exc:  # noqa: BLE001
        try:
            s.rollback()
        except Exception:
            pass
        return JSONResponse({"ok": False, "error": str(exc)[:200]}, status_code=500)
    finally:
        try:
            cm.__exit__(None, None, None)
        except Exception:
            pass


@doch_zak_tab_router.post("/app/dochazka-zak-tab/save-doch-meta")
async def dochazka_zak_tab_save_doch_meta(req: Request) -> JSONResponse:
    """Uloží k docházkovému záznamu (att_entry) POZNÁMKY a zaškrtávátka, které
    NEmají vliv na mzdy: note (zaměstnanec pozn.), vedouci_poznamka, ved_schvaleno.
    Zapisuje na AKTIVNÍ řádek — když byl původní opraven přes fix/entry (supersede),
    najde nový řádek (source_id=orig). Časy/hodiny se řeší přes fix/entry zvlášť."""
    from modules.erp.api.router import _uid_from_token_or_cookie
    uid = _uid_from_token_or_cookie(req)
    if not uid:
        return JSONResponse({"ok": False, "error": "unauthorized"}, status_code=401)
    if not _dzt_can(uid):
        return JSONResponse({"ok": False, "error": "forbidden"}, status_code=403)
    try:
        body = await req.json()
    except Exception:
        body = {}
    try:
        orig = int((body or {}).get("id"))
    except Exception:
        return JSONResponse({"ok": False, "error": "chybí id"}, status_code=400)
    has_note = "note" in (body or {})
    has_vp = "vedouci_poznamka" in (body or {})
    has_vs = "ved_schvaleno" in (body or {})
    note = str((body or {}).get("note") or "")[:2000] if has_note else None
    vp = str((body or {}).get("vedouci_poznamka") or "")[:2000] if has_vp else None
    vs = bool((body or {}).get("ved_schvaleno")) if has_vs else None
    from sqlalchemy import text as _t
    from modules.strategie_pg.application import service as _pg
    cm = _pg.get_session()
    s = cm.__enter__()
    try:
        # aktivní řádek: nový (source_id=orig) má přednost před orig samotným
        aid = s.execute(_t(
            "SELECT id FROM tenant.att_entry "
            "WHERE tenant_id=2 AND COALESCE(status,'')<>'superseded' AND (source_id=:o OR id=:o) "
            "ORDER BY (id=:o)::int ASC, id DESC LIMIT 1"), {"o": orig}).scalar()
        if not aid:
            return JSONResponse({"ok": False, "error": "Záznam nenalezen (nebo byl nahrazen)."},
                                status_code=404)
        sets, params = [], {"id": aid}
        if has_note:
            sets.append("note=:note"); params["note"] = note
        if has_vp:
            sets.append("vedouci_poznamka=:vp"); params["vp"] = vp
        if has_vs:
            sets.append("ved_schvaleno=:vs"); params["vs"] = vs
        if not sets:
            return JSONResponse({"ok": True, "id": aid, "note": "nic ke změně"})
        s.execute(_t("UPDATE tenant.att_entry SET " + ", ".join(sets) + " WHERE id=:id AND tenant_id=2"),
                  params)
        s.commit()
        return JSONResponse({"ok": True, "id": aid})
    except Exception as exc:  # noqa: BLE001
        try:
            s.rollback()
        except Exception:
            pass
        return JSONResponse({"ok": False, "error": str(exc)[:200]}, status_code=500)
    finally:
        try:
            cm.__exit__(None, None, None)
        except Exception:
            pass


@doch_zak_tab_router.post("/app/dochazka-zak-tab/widths")
async def dochazka_zak_tab_widths_set(req: Request) -> JSONResponse:
    """Uloží OSOBNÍ šířky uživatele do DB (kdokoliv z povolených 11). Tím je Claude
    může přečíst a povýšit na sdílené výchozí (kod bez _u<uid>)."""
    from modules.erp.api.router import _uid_from_token_or_cookie
    uid = _uid_from_token_or_cookie(req)
    if not uid:
        return JSONResponse({"ok": False, "error": "unauthorized"}, status_code=401)
    if not _dzt_can(uid):
        return JSONResponse({"ok": False, "error": "forbidden"}, status_code=403)
    try:
        body = await req.json()
    except Exception:
        body = {}
    widths = (body or {}).get("widths")
    if not isinstance(widths, dict):
        return JSONResponse({"ok": False, "error": "bad_widths"}, status_code=400)
    # sanitizace: jen {sloupec: kladné číslo}
    clean = {}
    for k, v in widths.items():
        try:
            iv = int(v)
        except Exception:
            continue
        if isinstance(k, str) and 20 <= iv <= 1200:
            clean[k[:40]] = iv
    import json as _j
    from sqlalchemy import text as _t
    from modules.strategie_pg.application import service as _pg
    key = _DZT_WIDTHS_KEY + "_u" + str(int(uid))
    cm = _pg.get_session()
    s = cm.__enter__()
    try:
        s.execute(_t(
            "INSERT INTO tenant.att_ui_pref (kod, hodnota, updated_by, updated_at) "
            "VALUES (:k, CAST(:v AS jsonb), :u, now()) "
            "ON CONFLICT (kod) DO UPDATE SET hodnota=CAST(:v AS jsonb), updated_by=:u, updated_at=now()"),
            {"k": key, "v": _j.dumps(clean), "u": int(uid)})
        s.commit()
        return JSONResponse({"ok": True, "ulozeno": len(clean)})
    except Exception as exc:  # noqa: BLE001
        try:
            s.rollback()
        except Exception:
            pass
        return JSONResponse({"ok": False, "error": str(exc)[:200]}, status_code=500)
    finally:
        try:
            cm.__exit__(None, None, None)
        except Exception:
            pass


@doch_zak_tab_router.post("/app/dochazka-zak-tab/delete-usek")
async def dochazka_zak_tab_delete_usek(req: Request) -> JSONResponse:
    """Smaže JEDEN výrobní úsek (řádek tenant.vyroba_work, kind='W') — zneaktivní ho
    (is_active=false) + audit do poznámky (kdo/proč). Postup dle Jirky (24.7.2026,
    ověřeno v kódu): work_alloc NEMAZAT (osiřel by řádek), sync is_active nevrací,
    v přehledech (Docházka new i Dušanův) je is_active filtrované → úsek zmizí, ale
    v DB zůstane pro dohledatelnost. Vratné (is_active zpět na true). Povinný důvod."""
    from modules.erp.api.router import _uid_from_token_or_cookie
    uid = _uid_from_token_or_cookie(req)
    if not uid:
        return JSONResponse({"ok": False, "error": "unauthorized"}, status_code=401)
    if not _dzt_can(uid):
        return JSONResponse({"ok": False, "error": "forbidden"}, status_code=403)
    try:
        body = await req.json()
    except Exception:
        body = {}
    try:
        rid = int((body or {}).get("id"))
    except Exception:
        return JSONResponse({"ok": False, "error": "chybí id řádku"}, status_code=400)
    reason = str((body or {}).get("reason") or "").strip()[:300]
    if not reason:
        return JSONResponse({"ok": False, "error": "Důvod smazání je povinný (kvůli auditu)."}, status_code=400)
    from sqlalchemy import text as _t
    from modules.strategie_pg.application import service as _pg
    cm = _pg.get_session()
    s = cm.__enter__()
    try:
        row = s.execute(_t("SELECT is_active, source_system FROM tenant.vyroba_work "
                           "WHERE id=:id AND tenant_id=2"), {"id": rid}).mappings().first()
        if not row:
            return JSONResponse({"ok": False, "error": "Úsek nenalezen."}, status_code=404)
        if row["is_active"] is False:
            return JSONResponse({"ok": True, "id": rid, "note": "už bylo smazané"})
        actor = s.execute(_t(
            "SELECT COALESCE(NULLIF(TRIM(COALESCE(first_name,'')||' '||COALESCE(last_name,'')),''),'?') "
            "FROM public.users WHERE id=:u"), {"u": int(uid)}).scalar() or "?"
        tag = "🗑 SMAZÁNO (" + actor + "): " + reason
        s.execute(_t(
            "UPDATE tenant.vyroba_work SET is_active=false, updated_at=now(), "
            " poznamka = CASE WHEN COALESCE(poznamka,'')='' THEN :tag ELSE poznamka || ' / ' || :tag END "
            "WHERE id=:id AND tenant_id=2 AND is_active=true"), {"id": rid, "tag": tag})
        s.commit()
        return JSONResponse({"ok": True, "id": rid})
    except Exception as exc:  # noqa: BLE001
        try:
            s.rollback()
        except Exception:
            pass
        return JSONResponse({"ok": False, "error": str(exc)[:200]}, status_code=500)
    finally:
        try:
            cm.__exit__(None, None, None)
        except Exception:
            pass


# ── Import z výkazu práce (EUROSOFT Work Report .xlsx) ────────────────────────
# List „Stunden": C2=pracovník (jen zobrazení), G5=číslo zam (podle něj se páruje),
# G2=výchozí zakázka, C5=výchozí druh činnosti. Datové řádky 8..34:
#   B=datum, C=hodiny (NETTO, pauza už odečtená), D=od, E=do, F=pauza,
#   G=činnost (název), H=činnost č. (= vyroba_cinnost.ec_cislo), I=poznámka, J=zakázka.
# Zapisujeme do tenant.vyroba_work stejně jako „Nový" (save-new), source_system='app'
# (jinak by řádky nebyly v přehledu — dataset filtruje source_system IN ('app','centrala1')).
# Ochrana proti duplicitě: (user_id, datum, od) — opakovaný import téhož výkazu nepřidá hodiny.
def _dzt_norm_date(v) -> str | None:
    """Různé podoby datumu z Excelu → 'YYYY-MM-DD' (None, když prázdné/nečitelné)."""
    import datetime as _dt
    if v is None:
        return None
    if isinstance(v, (_dt.datetime, _dt.date)):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d.%m.%y", "%Y/%m/%d"):
        try:
            return _dt.datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except Exception:
            pass
    return s[:10]


def _dzt_norm_time(v) -> str | None:
    """'HH:MM' z time/datetime/textu (None, když prázdné)."""
    import datetime as _dt
    if v is None:
        return None
    if isinstance(v, (_dt.time, _dt.datetime)):
        return v.strftime("%H:%M")
    s = str(v).strip()
    if not s:
        return None
    try:
        p = s.split(":")
        return "%02d:%02d" % (int(p[0]), int(p[1]))
    except Exception:
        return s[:5]


def _dzt_parse_vykaz(raw: bytes) -> tuple[dict, list[dict]]:
    """Naparsuje jeden Work Report (.xlsx) → (hlavička, řádky). Bez DB."""
    import io
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
    ws = wb["Stunden"] if "Stunden" in wb.sheetnames else wb.worksheets[0]

    def _c(ref):
        return ws[ref].value
    hdr = {
        "worker_name": (str(_c("C2")).strip() if _c("C2") is not None else ""),
        "cislo_zam": (str(_c("G5")).strip() if _c("G5") is not None else ""),
        "order_default": (str(_c("G2")).strip() if _c("G2") is not None else ""),
        "customer": (str(_c("C4")).strip() if _c("C4") is not None else ""),
    }
    rows: list[dict] = []
    for r in range(8, 35):
        datum = _dzt_norm_date(ws.cell(r, 2).value)   # B
        if not datum:
            continue
        hraw = ws.cell(r, 3).value                     # C = hodiny (netto)
        try:
            hod = float(hraw) if hraw not in (None, "") else 0.0
        except Exception:
            hod = 0.0
        if hod == 0:
            continue
        cin_cislo = ws.cell(r, 8).value                # H
        try:
            cin_cislo = int(cin_cislo) if cin_cislo not in (None, "") else None
        except Exception:
            cin_cislo = None
        zak = ws.cell(r, 10).value                      # J
        zak = (str(zak).strip() if zak not in (None, "") else "") or hdr["order_default"]
        pozn = ws.cell(r, 9).value                       # I
        rows.append({
            "datum": datum,
            "hodiny": round(hod, 2),
            "od": _dzt_norm_time(ws.cell(r, 4).value),   # D
            "do": _dzt_norm_time(ws.cell(r, 5).value),   # E
            "cin_name": (str(ws.cell(r, 7).value).strip() if ws.cell(r, 7).value else ""),  # G
            "cin_cislo": cin_cislo,
            "zakazka": zak,
            "poznamka": (str(pozn).strip() if pozn not in (None, "") else ""),
        })
    try:
        wb.close()
    except Exception:
        pass
    return hdr, rows


def _dzt_parse_one(filename: str, raw: bytes) -> dict:
    """(název, bajty) → {filename, hdr, rows} nebo {filename, error}. Bez DB."""
    if not raw:
        return {"filename": filename, "error": "Prázdný soubor."}
    if len(raw) > 6 * 1024 * 1024:
        return {"filename": filename, "error": "Soubor je příliš velký (max 6 MB)."}
    try:
        hdr, rows = _dzt_parse_vykaz(raw)
    except Exception as exc:  # noqa: BLE001
        return {"filename": filename, "error": "Nepodařilo se přečíst Excel: " + str(exc)[:120]}
    return {"filename": filename, "hdr": hdr, "rows": rows}


def _dzt_day_segments(od_dt, do_dt, net_h: float, break_h: float):
    """Rozseká den [od,do] na úseky jako mobilní „Makat": práce/pauza/práce.
    Bez pauzy → jeden pracovní úsek. S pauzou → práce (půl čisté) + pauza + práce
    (druhá půl); součet práce = net_h, pauza uprostřed. Vrací [(druh, start, end)]."""
    from datetime import timedelta as _td
    if not break_h or break_h <= 0.011:
        return [("work", od_dt, do_dt)]
    w1_end = od_dt + _td(hours=net_h / 2.0)
    brk_end = w1_end + _td(hours=break_h)
    return [("work", od_dt, w1_end), ("break", w1_end, brk_end), ("work", brk_end, do_dt)]


def _dzt_process_parsed(parsed: list[dict], do_commit: bool, uid: int, s) -> tuple[list, dict, dict]:
    """Náhled/zápis importu. Zapisuje STEJNĚ jako mobilní „Makat" / „Přidat záznam":
    přítomnost do tenant.att_entry (práce + pauza → mzda) A úsek do tenant.work_alloc
    (→ vyroba_work → zakázky). Pauza se založí jako samostatný záznam, takže čistá
    práce sedí (06:00–14:30 s pauzou 0,5 = 8 h práce + 0,5 h pauza). NEcommituje.
    Pojistky (jako fix/add): uzamčený měsíc a překryv s existující docházkou → přeskočí
    (chrání před dvojí mzdou). Vrací (out_files, totals, post_actions)."""
    import datetime as _dtm
    from sqlalchemy import text as _t
    # činnost: ec_cislo→(id,name,icon) + name(lower)→(id,name,icon)
    cin_by_ec: dict = {}
    cin_by_name: dict = {}
    for cid, ec, nm, ic in s.execute(_t(
            "SELECT id, ec_cislo, name, COALESCE(icon,'') FROM tenant.vyroba_cinnost "
            "WHERE COALESCE(active,true)")).all():
        rec = (int(cid), nm, ic or None)
        if ec is not None:
            try:
                cin_by_ec[int(ec)] = rec
            except Exception:
                pass
        if nm:
            cin_by_name[str(nm).strip().lower()] = rec
    # typy záznamů
    work_tid = s.execute(_t("SELECT id FROM tenant.att_entry_type WHERE tenant_id=2 AND code='work'")).scalar()
    break_tid = s.execute(_t("SELECT id FROM tenant.att_entry_type WHERE tenant_id=2 AND code='break'")).scalar()

    from modules.erp.api.router import _att_period_locked, _att_fix_overlap  # guardy (mzdy)

    out_files = []
    total = {"ok": 0, "duplicate": 0, "error": 0, "inserted": 0}
    emp_days: set = set()        # (employee_id, 'YYYY-MM-DD') pro přepočet fondu
    dmin = dmax = None           # rozsah dní pro fold work_alloc→vyroba_work
    for pf in parsed:
        if pf.get("error"):
            out_files.append({"filename": pf["filename"], "error": pf["error"]})
            total["error"] += 1
            continue
        hdr, rows = pf["hdr"], pf["rows"]
        cz = str(hdr.get("cislo_zam") or "").strip()
        emp_id = emp_uid = emp_name = None
        if cz:
            res = s.execute(_t(
                "SELECT id, user_id, full_name FROM tenant.att_employee "
                "WHERE tenant_id=2 AND cislo_zam=:cz AND user_id IS NOT NULL LIMIT 1"),
                {"cz": cz}).first()
            if res:
                emp_id, emp_uid, emp_name = res[0], res[1], res[2]
        worker_ok = emp_id is not None

        out_rows = []
        f_sum = {"ok": 0, "duplicate": 0, "error": 0}
        for r in rows:
            warn = []
            status = "ok"
            # činnost
            cin_id = cin_db = cin_icon = None
            if r["cin_cislo"] is not None and r["cin_cislo"] in cin_by_ec:
                cin_id, cin_db, cin_icon = cin_by_ec[r["cin_cislo"]]
            elif r["cin_name"] and r["cin_name"].lower() in cin_by_name:
                cin_id, cin_db, cin_icon = cin_by_name[r["cin_name"].lower()]
            # zakázka
            zak = (r["zakazka"] or "").strip() or None
            is_rezie = bool(zak and zak.strip().lower() in ("rezie", "režie"))
            zak_ok = True
            if zak and not is_rezie:
                zr = s.execute(_t(
                    "SELECT COALESCE(pichatelna,false), COALESCE(nazev,'') FROM tenant.zakazka "
                    "WHERE tenant_id=2 AND cislo=:z LIMIT 1"), {"z": zak}).first()
                if not zr:
                    zak_ok = False
                    warn.append("zakázka '" + zak + "' neexistuje")
                elif not zr[0]:
                    zak_ok = False
                    warn.append("zakázka '" + zak + "' není píchatelná")
            zak_nazev = None
            if zak and not is_rezie and zak_ok:
                zak_nazev = s.execute(_t("SELECT nazev FROM tenant.zakazka WHERE tenant_id=2 AND cislo=:c"),
                                      {"c": zak}).scalar()
            # časy + pauza
            od_dt = kon_dt = None
            net_h = float(r["hodiny"] or 0)
            break_h = 0.0
            if r["od"] and r["do"]:
                try:
                    od_dt = _dtm.datetime.strptime(r["datum"] + " " + r["od"], "%Y-%m-%d %H:%M")
                    kon_dt = _dtm.datetime.strptime(r["datum"] + " " + r["do"], "%Y-%m-%d %H:%M")
                except Exception:
                    od_dt = kon_dt = None
            gross = (kon_dt - od_dt).total_seconds() / 3600.0 if (od_dt and kon_dt) else None
            if gross is not None:
                break_h = round(gross - net_h, 2)
            # rozhodnutí o stavu (money-safe)
            if not worker_ok:
                status = "error"; warn.insert(0, "pracovník č. " + (cz or "?") + " nenalezen")
            elif not (od_dt and kon_dt):
                status = "error"; warn.insert(0, "chybí čas od/do")
            elif kon_dt <= od_dt:
                status = "error"; warn.insert(0, "konec není po začátku")
            elif break_h < -0.02:
                status = "error"; warn.insert(0, "hodiny > rozpětí od–do (zkontroluj výkaz)")
            elif zak and not is_rezie and not zak_ok:
                status = "error"
            else:
                if cin_id is None:
                    warn.append("činnost nerozpoznána — přítomnost se založí, ale na zakázku se nepromítne")
                day_d = _dtm.date.fromisoformat(r["datum"])
                if _att_period_locked(s, day_d):
                    status = "error"; warn.insert(0, "měsíc je uzamčen (mzdy zpracovány)")
                else:
                    ovl = _att_fix_overlap(s, emp_id, od_dt, kon_dt, None)
                    if ovl:
                        status = "duplicate"; warn.insert(0, "překryv s docházkou " + str(ovl) + " — přeskočí se")
            if break_h < 0:
                break_h = 0.0
            out_rows.append({
                "datum": r["datum"], "od": r["od"], "do": r["do"], "hodiny": round(net_h, 2),
                "pauza": break_h, "cin_cislo": r["cin_cislo"], "cin_name": r["cin_name"],
                "cin_id": cin_id, "cin_db": cin_db, "zakazka": zak, "zakazka_ok": zak_ok,
                "poznamka": r["poznamka"], "status": status, "warn": warn,
                "_od": od_dt, "_kon": kon_dt, "_net": net_h, "_brk": break_h,
                "_cin": cin_id, "_cinn": cin_db, "_cinic": cin_icon,
                "_zak": (None if is_rezie else zak), "_zaknz": zak_nazev, "_rez": is_rezie,
            })
            f_sum[status] = f_sum.get(status, 0) + 1

        inserted_ids = []
        if do_commit and worker_ok:
            note = "IMPORT výkaz práce"
            for orow in out_rows:
                if orow["status"] != "ok":
                    continue
                try:
                    day_iso = orow["datum"]
                    for druh, sdt, edt in _dzt_day_segments(orow["_od"], orow["_kon"], orow["_net"], orow["_brk"]):
                        seg_h = round((edt - sdt).total_seconds() / 3600.0, 2)
                        tid = work_tid if druh == "work" else break_tid
                        pr = (orow["_zak"] if druh == "work" else None)
                        aid = s.execute(_t(
                            "INSERT INTO tenant.att_entry (tenant_id, employee_id, entry_date, entry_type_id, hours, "
                            "started_at, ended_at, project_ref, note, status, source, is_active, "
                            "created_by_id, created_at, updated_at) "
                            "VALUES (2,:e,:d,:ti,:h,CAST(:ns AS timestamp),CAST(:ne AS timestamp),:pr,:n,"
                            "'approved','import',false,:u,now(),now()) RETURNING id"),
                            {"e": emp_id, "d": day_iso, "ti": tid, "h": seg_h,
                             "ns": sdt.isoformat(sep=" "), "ne": edt.isoformat(sep=" "),
                             "pr": pr, "n": (note if druh == "work" else note + " (pauza)"), "u": uid}).scalar()
                        inserted_ids.append(int(aid))
                        if druh == "work":
                            # C24 29.7.2026 (krok 5-kód): import píše úsek přímo do JEDNÉ
                            # tabulky vyroba_work (ne work_alloc). Vazba att_entry_id=aid
                            # (čistá, přes právě založený att_entry). Režie = zakazka_ref='Rezie'
                            # (žádný is_rezie). Denorm názvy se neukládají (join v přehledech).
                            s.execute(_t(
                                "INSERT INTO tenant.vyroba_work (tenant_id,user_id,cislo_zam,datum,od,konec,"
                                "zakazka_ref,cinnost_id,hodiny,source_system,att_entry_id,created_at,updated_at) "
                                "VALUES (2,:u,"
                                "(SELECT cislo_zam FROM tenant.att_employee e WHERE e.user_id=:u AND e.tenant_id=2 AND e.cislo_zam IS NOT NULL LIMIT 1),"
                                ":d,CAST(:ns AS timestamptz),CAST(:ne AS timestamptz),"
                                ":pr,(SELECT id FROM tenant.vyroba_cinnost WHERE id=:ci LIMIT 1),:h,'import',:aid,now(),now())"),
                                {"u": emp_uid, "d": day_iso, "ns": sdt.isoformat(sep=" "), "ne": edt.isoformat(sep=" "),
                                 "pr": (orow["_zak"] or (_REZIE_REF if orow["_rez"] else None)),
                                 "ci": orow["_cin"], "h": seg_h, "aid": int(aid)})
                    emp_days.add((emp_id, day_iso))
                    if dmin is None or day_iso < dmin:
                        dmin = day_iso
                    if dmax is None or day_iso > dmax:
                        dmax = day_iso
                    orow["status"] = "inserted"
                except Exception as exc:  # noqa: BLE001
                    orow["status"] = "error"
                    orow["warn"].append("zápis selhal: " + str(exc)[:80])
                    f_sum["error"] = f_sum.get("error", 0) + 1

            # nový součet po zápisu
            f_sum = {"ok": 0, "duplicate": 0, "error": 0}
            for orow in out_rows:
                st = orow["status"] if orow["status"] in ("duplicate", "error") else ("inserted" if orow["status"] == "inserted" else "ok")
                f_sum[st if st in f_sum else "ok"] = f_sum.get(st if st in f_sum else "ok", 0) + 1

        for orow in out_rows:
            for k in ("_od", "_kon", "_net", "_brk", "_cin", "_cinn", "_cinic", "_zak", "_zaknz", "_rez"):
                orow.pop(k, None)

        total["ok"] += sum(1 for o in out_rows if o["status"] == "ok")
        total["duplicate"] += sum(1 for o in out_rows if o["status"] == "duplicate")
        total["error"] += sum(1 for o in out_rows if o["status"] == "error")
        total["inserted"] += sum(1 for o in out_rows if o["status"] == "inserted")
        out_files.append({
            "filename": pf["filename"],
            "worker": {"cislo": cz, "name_excel": hdr.get("worker_name"),
                       "name_db": emp_name, "resolved": worker_ok},
            "customer": hdr.get("customer"),
            "order_default": hdr.get("order_default"),
            "rows": out_rows,
            "summary": f_sum,
            "inserted_ids": inserted_ids,
        })

    return out_files, total, {"emp_days": emp_days, "frm": dmin, "to": dmax}


def _dzt_import_run(parsed: list[dict], do_commit: bool, uid: int) -> JSONResponse:
    """Otevře PG session, spustí _dzt_process_parsed, commitne a vrátí odpověď.
    Po zápisu přepočítá fond dotčených dní. Úseky se píší rovnou do vyroba_work
    (C24 29.7.2026, krok 5/7 — už žádný fold z work_alloc)."""
    from modules.strategie_pg.application import service as _pg
    cm = _pg.get_session()
    s = cm.__enter__()
    post = {}
    try:
        out_files, total, post = _dzt_process_parsed(parsed, do_commit, uid, s)
        if do_commit:
            s.commit()
    except Exception as exc:  # noqa: BLE001
        try:
            s.rollback()
        except Exception:
            pass
        try:
            cm.__exit__(None, None, None)
        except Exception:
            pass
        return JSONResponse({"ok": False, "error": str(exc)[:200]}, status_code=500)
    finally_done = False
    try:
        cm.__exit__(None, None, None)
        finally_done = True
    except Exception:
        pass
    # po-commit akce (vlastní session uvnitř helperů) — jen při ostrém zápisu
    if do_commit and post:
        try:
            from modules.erp.api.router import _att_automat_recalc_day
            import datetime as _d2
            for emp_id, day_iso in (post.get("emp_days") or set()):
                try:
                    _att_automat_recalc_day(emp_id, _d2.date.fromisoformat(day_iso))
                except Exception:
                    pass
        except Exception:
            pass
        # C24 29.7.2026 (krok 7): app fold zrušen — import píše úseky rovnou do vyroba_work
        # (viz INSERT výše), takže není co přelévat z work_alloc.
    return JSONResponse({"ok": True, "mode": ("commit" if do_commit else "preview"),
                         "files": out_files, "totals": total})


# ── Sdílená složka výkazů (Dušan) — čtení přes EUROSOFT MCP ───────────────────
# UNC \\192.168.30.11\Data\... = lokální D:\data\... na EC-SERVER2, což je
# whitelistovaný RO root MCP (MCP_FS_RO_ROOTS) — jako bank_api / cenik_engine.
_DZT_IMPORT_UNC = r"\\192.168.30.11\Data\Vedouci vyroby\ImportDochazky"
_DZT_IMPORT_LOCAL = r"D:\data\Vedouci vyroby\ImportDochazky"


def _dzt_mcp():
    from modules.conversation.application.eurosoft_mcp_client import get_eurosoft_mcp_client
    return get_eurosoft_mcp_client()


def _dzt_slozka_list() -> tuple[list[dict] | None, str | None]:
    """Vypíše .xlsx ve složce výkazů přes MCP. Vrací (soubory, chyba)."""
    import json as _j
    mcp = _dzt_mcp()
    if mcp is None:
        return None, "EUROSOFT MCP je nedostupný (spojení s on-prem serverem)."
    try:
        raw = mcp.call_tool_sync("eurosoft_eurosoft_file_list",
                                 {"user_namespace": "ro", "base_override": _DZT_IMPORT_LOCAL, "subpath": ""},
                                 conversation_id=None)
    except Exception as exc:  # noqa: BLE001
        return None, "Nepodařilo se načíst složku: " + str(exc)[:150]
    r = _j.loads(raw) if isinstance(raw, str) else raw
    if not isinstance(r, dict) or r.get("ok") is False:
        return None, str((r or {}).get("error") or "Složku se nepodařilo otevřít.")[:200]
    out = []
    for it in (r.get("items") or []):
        if it.get("type") == "file" and str(it.get("name", "")).lower().endswith(".xlsx"):
            out.append({"name": it["name"], "size": it.get("size"), "modified": it.get("modified")})
    return out, None


def _dzt_slozka_read(filename: str) -> tuple[bytes | None, str | None]:
    """Přečte jeden .xlsx ze složky výkazů přes MCP (base64 → bytes)."""
    import base64 as _b64
    import json as _j
    mcp = _dzt_mcp()
    if mcp is None:
        return None, "EUROSOFT MCP je nedostupný."
    fn = (filename or "").replace("\\", "/").split("/")[-1].strip()  # jen holé jméno
    if not fn or not fn.lower().endswith(".xlsx"):
        return None, "Neplatný soubor."
    try:
        raw = mcp.call_tool_sync("eurosoft_eurosoft_file_read",
                                 {"user_namespace": "ro", "base_override": _DZT_IMPORT_LOCAL,
                                  "path": fn, "encoding": "base64"}, conversation_id=None)
    except Exception as exc:  # noqa: BLE001
        return None, "Čtení selhalo: " + str(exc)[:150]
    r = _j.loads(raw) if isinstance(raw, str) else raw
    if isinstance(r, dict) and r.get("ok") is False:
        return None, str(r.get("error") or "Soubor se nepodařilo přečíst.")[:200]
    b64 = (r.get("content") or r.get("data") or "") if isinstance(r, dict) else str(r)
    try:
        return _b64.b64decode(b64), None
    except Exception:
        return None, "Obsah souboru se nepodařilo dekódovat."


@doch_zak_tab_router.post("/app/dochazka-zak-tab/import-vykaz")
async def dochazka_zak_tab_import_vykaz(
    req: Request,
    files: list[UploadFile] = File(...),
    commit: str = Form("0"),
) -> JSONResponse:
    """Import NAHRANÝCH výkazů (multipart). commit=0 náhled / 1 zápis."""
    from modules.erp.api.router import _uid_from_token_or_cookie
    uid = _uid_from_token_or_cookie(req)
    if not uid:
        return JSONResponse({"ok": False, "error": "unauthorized"}, status_code=401)
    if not _dzt_can(uid):
        return JSONResponse({"ok": False, "error": "forbidden"}, status_code=403)
    do_commit = str(commit or "0").strip() in ("1", "true", "yes", "on")
    if not files:
        return JSONResponse({"ok": False, "error": "Nebyl nahrán žádný soubor."}, status_code=400)
    parsed = []
    for f in files:
        try:
            raw = await f.read()
        except Exception:
            raw = b""
        parsed.append(_dzt_parse_one(f.filename, raw))
    return _dzt_import_run(parsed, do_commit, uid)


@doch_zak_tab_router.get("/app/dochazka-zak-tab/import-slozka")
def dochazka_zak_tab_import_slozka(req: Request) -> JSONResponse:
    """Seznam .xlsx ve sdílené složce výkazů (Dušanova ImportDochazky) přes MCP."""
    from modules.erp.api.router import _uid_from_token_or_cookie
    uid = _uid_from_token_or_cookie(req)
    if not uid:
        return JSONResponse({"ok": False, "error": "unauthorized"}, status_code=401)
    if not _dzt_can(uid):
        return JSONResponse({"ok": False, "error": "forbidden"}, status_code=403)
    files, err = _dzt_slozka_list()
    if err:
        return JSONResponse({"ok": False, "error": err, "unc": _DZT_IMPORT_UNC})
    return JSONResponse({"ok": True, "unc": _DZT_IMPORT_UNC, "files": files})


@doch_zak_tab_router.post("/app/dochazka-zak-tab/import-vykaz-slozka")
async def dochazka_zak_tab_import_vykaz_slozka(req: Request) -> JSONResponse:
    """Import VYBRANÝCH souborů ze sdílené složky (čte se přes MCP). commit=0/1."""
    from modules.erp.api.router import _uid_from_token_or_cookie
    uid = _uid_from_token_or_cookie(req)
    if not uid:
        return JSONResponse({"ok": False, "error": "unauthorized"}, status_code=401)
    if not _dzt_can(uid):
        return JSONResponse({"ok": False, "error": "forbidden"}, status_code=403)
    try:
        body = await req.json()
    except Exception:
        body = {}
    names = [str(x) for x in ((body or {}).get("files") or []) if x]
    do_commit = str((body or {}).get("commit") or "0").strip() in ("1", "true", "yes", "on")
    if not names:
        return JSONResponse({"ok": False, "error": "Nevybral jsi žádný soubor."}, status_code=400)
    parsed = []
    for nm in names[:50]:
        raw, err = _dzt_slozka_read(nm)
        if err:
            parsed.append({"filename": nm, "error": err})
        else:
            parsed.append(_dzt_parse_one(nm, raw))
    return _dzt_import_run(parsed, do_commit, uid)
