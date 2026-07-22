"""
Landmark – měsíční podklad pro fakturaci (mzdy).

Vezme z vygenerovaných výplatnic (`tenant.payslip_item`) náhradu za oblečení
(složka 794) a náhradu HO (složka 795) po lidech, za obě firmy (EC/ES),
sestaví Excel a pošle ho e-mailem na nakup@eurosoft.com.

Fakturace Landmarku = 9,06 % × (oblečení + HO) + 21 % DPH (ověřeno na dubnu
a květnu 2026, obě firmy). Podrobně: docs/Z_landmark_podklad_mzdy.md.

Spouštění:
  - automaticky vždy 15. v měsíci (odešle podklad za PŘEDCHOZÍ měsíc) – scheduler
    `landmark_sched_start()` (jen primár, registruje se v lifespanu),
  - ručně / test přes endpoint GET /app/mzdy/landmark-send?rok=&mesic=[&to=].

Peta + Claude (ID26), 22. 7. 2026.
"""
import io
import os
import logging
import datetime as _dt
import tempfile

from fastapi import APIRouter, Request
from fastapi.responses import JSONResponse
from sqlalchemy import text as _t

logger = logging.getLogger(__name__)

landmark_router = APIRouter(prefix="/api/v1/erp", tags=["landmark"])

LANDMARK_TO = "nakup@eurosoft.com"
RATE = 0.0906
_MES_NAZEV = {1: "lednové", 2: "únorové", 3: "březnové", 4: "dubnové", 5: "květnové",
              6: "červnové", 7: "červencové", 8: "srpnové", 9: "zářijové",
              10: "říjnové", 11: "listopadové", 12: "prosincové"}

# stav scheduleru (proces-lokální, jako mirror_sched)
_LM_TASK = [None]
_LM_STOP = [False]


# ───────────────────────────── data ─────────────────────────────

def _fetch(rok: int, mesic: int) -> dict:
    """Vrátí {'EC': [(cislo, jmeno, obl, ho), ...], 'ES': [...]} – jen lidi s obl/ho > 0."""
    from core.database_data import get_data_session as _g
    s = _g()
    try:
        rows = s.execute(_t(
            "SELECT c.code AS firma, em.cislo_zam AS cislo, em.full_name AS jmeno, "
            "  COALESCE(SUM(p.koruny) FILTER (WHERE p.cislo_ms = 794), 0) AS obl, "
            "  COALESCE(SUM(p.koruny) FILTER (WHERE p.cislo_ms = 795), 0) AS ho "
            "FROM tenant.payslip_item p "
            "LEFT JOIN tenant.company c       ON c.id = p.company_id "
            "LEFT JOIN tenant.att_employee em ON em.id = p.employee_id "
            "WHERE p.tenant_id = 2 AND p.rok = :y AND p.mesic = :m "
            "  AND p.cislo_ms IN (794, 795) "
            "GROUP BY c.code, em.cislo_zam, em.full_name "
            "ORDER BY firma, jmeno"),
            {"y": rok, "m": mesic}).fetchall()
    finally:
        s.close()
    data = {"EC": [], "ES": []}
    for firma, cislo, jmeno, obl, ho in rows:
        obl = float(obl or 0)
        ho = float(ho or 0)
        if obl == 0 and ho == 0:
            continue
        if firma in data:
            data[firma].append((cislo or "", jmeno or "", obl, ho))
    return data


# ───────────────────────────── Excel ─────────────────────────────

def _build_xlsx(rok: int, mesic: int, data: dict) -> bytes:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    F = "Arial"
    hdr_fill = PatternFill("solid", fgColor="1F4E78")
    hdr_font = Font(name=F, bold=True, color="FFFFFF", size=10)
    title_font = Font(name=F, bold=True, size=13)
    note_font = Font(name=F, italic=True, size=9, color="555555")
    sum_fill = PatternFill("solid", fgColor="DDEBF7")
    thin = Side(style="thin", color="BBBBBB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    kc = '#,##0 "Kč"'
    cols = ["Ev. číslo", "Jméno", "Náhrada za oblečení\n(Kč)", "Náhrada HO\n(Kč)", "Celkem optim.\n(Kč)"]
    nazev_mes = _MES_NAZEV.get(mesic, str(mesic))

    def sheet(ws, title, rows):
        ws.sheet_view.showGridLines = False
        ws["A1"] = title
        ws["A1"].font = title_font
        ws["A2"] = "%s mzdy %d – podklad pro LANDMARK TAX s.r.o." % (nazev_mes, rok)
        ws["A2"].font = note_font
        hr = 4
        for j, c in enumerate(cols, 1):
            cell = ws.cell(hr, j, c)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        r = hr + 1
        tobl = tho = 0.0
        for cislo, jmeno, obl, ho in rows:
            tobl += obl
            tho += ho
            ws.cell(r, 1, cislo)
            ws.cell(r, 2, jmeno)
            ws.cell(r, 3, obl if obl else None)
            ws.cell(r, 4, ho if ho else None)
            ws.cell(r, 5, obl + ho)
            for j in range(1, 6):
                cc = ws.cell(r, j)
                cc.border = border
                cc.font = Font(name=F, size=10)
            for j in (3, 4, 5):
                ws.cell(r, j).number_format = kc
            ws.cell(r, 1).alignment = Alignment(horizontal="center")
            r += 1
        ws.cell(r, 2, "CELKEM").font = Font(name=F, bold=True, size=10)
        for j, val in ((3, tobl), (4, tho), (5, tobl + tho)):
            cc = ws.cell(r, j, val)
            cc.number_format = kc
            cc.font = Font(name=F, bold=True, size=10)
        for j in range(1, 6):
            ws.cell(r, j).fill = sum_fill
            ws.cell(r, j).border = border
        opt = tobl + tho
        fee = round(opt * RATE, 2)
        dph = round(fee * 0.21, 2)
        cel = round(fee + dph)
        b = r + 2
        ws.cell(b, 2, "Odhad fakturace Landmark").font = Font(name=F, bold=True, size=11)
        for i, (lbl, val, fmt) in enumerate([
                ("Optimalizováno celkem (oblečení + HO)", opt, kc),
                ("Sazba Landmark", RATE, "0.00%"),
                ("Fakturace bez DPH", fee, kc),
                ("DPH 21 %", dph, kc),
                ("Celkem s DPH (odhad, po zaokrouhlení)", cel, kc)], 1):
            ws.cell(b + i, 2, lbl).font = Font(name=F, size=10)
            cc = ws.cell(b + i, 4, val)
            cc.number_format = fmt
            cc.font = Font(name=F, size=10, bold=lbl.startswith("Fakturace"))
        for j, w in enumerate([10, 26, 16, 13, 14], 1):
            ws.column_dimensions[get_column_letter(j)].width = w
        ws.row_dimensions[hr].height = 30
        return opt, fee, cel

    wb = openpyxl.Workbook()
    wsi = wb.active
    wsi.title = "Info"
    wsi.sheet_view.showGridLines = False
    wsi.column_dimensions["A"].width = 100
    info = [
        ("Podklad pro Landmark – %s mzdy %d" % (nazev_mes, rok), title_font),
        ("", None),
        ("Vytazeno primo ze systemu STRATEGIE (payslip_item). Slozky: 794 = obleceni, 795 = HO.", Font(name=F, size=10)),
        ("Fakturace = 9,06 % x (obleceni + HO), + 21 % DPH (overeno na dubnu a kvetnu 2026).", Font(name=F, size=10)),
        ("Souhrn pro Landmark je na konci zalozek EC a ES; detail po lidech pro kontrolu.", Font(name=F, size=10)),
    ]
    for i, (txt, ft) in enumerate(info, 1):
        cc = wsi.cell(i, 1, txt)
        if ft:
            cc.font = ft
    ws1 = wb.create_sheet("EC – EUROSOFT Control")
    o1, f1, c1 = sheet(ws1, "EUROSOFT – Control s.r.o.  (IČ 27960862)", data.get("EC", []))
    ws2 = wb.create_sheet("ES – EUROSOFT System")
    o2, f2, c2 = sheet(ws2, "EUROSOFT – System s.r.o.  (IČ 26411741)", data.get("ES", []))
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue(), (o1, f1, c1), (o2, f2, c2)


# ───────────────────────────── e-mail ─────────────────────────────

def _send_mail(to: str, subject: str, body: str, fname: str, content: bytes) -> None:
    """Pošle e-mail s XLSX přílohou z default persony (tenant 2) přes EWS."""
    from modules.notifications.application.email_service import (
        _get_account, _resolve_persona_email_creds, _get_default_persona_id,
    )
    from exchangelib import Message, Mailbox, FileAttachment

    # APID / readonly outbound guard (jako v send_email_or_raise)
    if (os.environ.get("STRATEGIE_ENV", "").lower() == "apid"
            or os.environ.get("STRATEGIE_READONLY_OUTBOUND") == "1"):
        logger.warning("[landmark] APID readonly – e-mail NEodeslan (to=%s)", to)
        return

    pid = _get_default_persona_id()
    creds = _resolve_persona_email_creds(pid, 2)
    if creds:
        account = _get_account(email=creds["email"], password=creds["password"], server=creds["server"])
    else:
        account = _get_account()
    to_list = [a.strip() for a in str(to).replace(";", ",").split(",") if a.strip()]
    msg = Message(
        account=account,
        subject=subject,
        body=body,
        to_recipients=[Mailbox(email_address=a) for a in to_list],
    )
    msg.attach(FileAttachment(name=fname, content=content))
    msg.send_and_save()
    logger.info("[landmark] mail odeslan | to=%s | subj=%s", to_list, subject)


# ───────────────────────── hlavní funkce ─────────────────────────

def build_and_send(rok: int, mesic: int, to: str = LANDMARK_TO) -> dict:
    """Sestaví podklad za (rok, mesic) a pošle mailem. Vrací souhrn."""
    data = _fetch(rok, mesic)
    n_ec = len(data.get("EC", []))
    n_es = len(data.get("ES", []))
    if n_ec == 0 and n_es == 0:
        raise RuntimeError("Za %d/%d nejsou v payslip_item zadne slozky 794/795 – neni co poslat "
                           "(uz probehla mzdova uzaverka @@MZDY?)." % (mesic, rok))
    content, ec_t, es_t = _build_xlsx(rok, mesic, data)
    nazev = _MES_NAZEV.get(mesic, str(mesic))
    fname = "Podklad_Landmark_%d_%02d.xlsx" % (rok, mesic)
    subject = "Podklad Landmark – %s mzdy %d (EUROSOFT EC + ES)" % (nazev, rok)
    body = (
        "Dobry den,\n\n"
        "v priloze posilam podklad pro Landmark za %s mzdy %d (EUROSOFT Control i System).\n\n"
        "Souhrn (obleceni + HO -> fakturace 9,06 %% + 21 %% DPH):\n"
        "  EC: optimalizovano %s Kc  ->  fakturace %s Kc bez DPH  ->  cca %s Kc s DPH  (%d lidi)\n"
        "  ES: optimalizovano %s Kc  ->  fakturace %s Kc bez DPH  ->  cca %s Kc s DPH  (%d lidi)\n\n"
        "Detail po lidech je v prilozenem Excelu (zalozky EC a ES).\n\n"
        "Automaticky generovano systemem STRATEGIE.\n"
        % (nazev, rok,
           _cz(ec_t[0]), _cz(ec_t[1]), _cz(ec_t[2]), n_ec,
           _cz(es_t[0]), _cz(es_t[1]), _cz(es_t[2]), n_es)
    )
    _send_mail(to, subject, body, fname, content)
    return {"rok": rok, "mesic": mesic, "to": to, "ec_lidi": n_ec, "es_lidi": n_es,
            "ec_fakturace": ec_t[1], "es_fakturace": es_t[1], "soubor": fname}


def _cz(x: float) -> str:
    x = float(x or 0)
    if x % 1:
        return "{:,.2f}".format(x).replace(",", " ").replace(".", ",")
    return "{:,.0f}".format(x).replace(",", " ")


def _prev_month(today: _dt.date):
    first = today.replace(day=1)
    last_prev = first - _dt.timedelta(days=1)
    return last_prev.year, last_prev.month


# ───────────────────────────── scheduler ─────────────────────────────

def _marker_path(rok: int, mesic: int) -> str:
    return os.path.join(tempfile.gettempdir(), "landmark_sent_%d_%02d.flag" % (rok, mesic))


def _landmark_tick():
    """Jednou denně: pokud je 15. a za předchozí měsíc jsme ještě neposlali, pošli."""
    today = _dt.date.today()
    if today.day != 15:
        return
    rok, mesic = _prev_month(today)
    mp = _marker_path(rok, mesic)
    if os.path.exists(mp):
        return
    try:
        res = build_and_send(rok, mesic, LANDMARK_TO)
        with open(mp, "w") as fh:
            fh.write(_dt.datetime.now().isoformat())
        logger.info("[landmark] automaticky odeslano za %d/%d: %s", mesic, rok, res)
    except Exception as exc:
        logger.warning("[landmark] automaticke odeslani za %d/%d SELHALO: %s", mesic, rok, exc)


async def _landmark_loop():
    import asyncio as _aio
    while not _LM_STOP[0]:
        try:
            await _aio.sleep(3600)  # každou hodinu zkontroluj datum
            loop = _aio.get_event_loop()
            await loop.run_in_executor(None, _landmark_tick)
        except _aio.CancelledError:
            break
        except Exception as e:
            logger.warning("[landmark_loop] %s", e)


def landmark_sched_start():
    import asyncio as _aio
    if _LM_TASK[0] is not None and not _LM_TASK[0].done():
        return
    _LM_STOP[0] = False
    try:
        _LM_TASK[0] = _aio.create_task(_landmark_loop())
        logger.info("[landmark] scheduler nastartovan (kontrola 1x/hod, odesila 15. v mesici)")
    except Exception as e:
        logger.warning("[landmark] scheduler start failed: %s", e)


def landmark_sched_stop_now():
    _LM_STOP[0] = True
    if _LM_TASK[0] is not None and not _LM_TASK[0].done():
        _LM_TASK[0].cancel()


# ───────────────────────────── endpoint (ruční / test) ─────────────────────────────

@landmark_router.get("/app/mzdy/landmark-send")
def landmark_send(req: Request):
    """Ruční/testovací odeslání podkladu. Parent-only.
    ?rok=&mesic= (default = předchozí měsíc), volitelně ?to= (default nakup@eurosoft.com)."""
    from modules.erp.api.router import _uid_from_token_or_cookie, _is_cockpit
    from core.database_data import get_data_session as _g
    uid = _uid_from_token_or_cookie(req)
    if not uid:
        return JSONResponse({"ok": False, "error": "forbidden"}, status_code=403)
    _s = _g()
    try:
        if not _is_cockpit(_s, uid):
            return JSONResponse({"ok": False, "error": "forbidden"}, status_code=403)
    finally:
        _s.close()
    today = _dt.date.today()
    dr, dm = _prev_month(today)
    try:
        rok = int(req.query_params.get("rok") or dr)
        mesic = int(req.query_params.get("mesic") or dm)
    except Exception:
        return JSONResponse({"ok": False, "error": "spatny rok/mesic"}, status_code=400)
    to = (req.query_params.get("to") or LANDMARK_TO).strip()
    try:
        res = build_and_send(rok, mesic, to)
        return JSONResponse({"ok": True, "poslano": res})
    except Exception as exc:
        logger.error("[landmark] ruční odeslání selhalo: %s", exc)
        return JSONResponse({"ok": False, "error": str(exc)}, status_code=500)
