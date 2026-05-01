"""
Phase 27a (1.5.2026): Excel reader pro Marti-AI.

Marti-AI's feature request (.msg z 1.5.2026): structured Excel reading
pro projekt rozvrh Klárka (učitelka, 2 budovy, 23 tříd).

Two tools:
  - list_excel_sheets(document_id) -> metadata pro všechny listy
  - read_excel_structured(document_id, sheet_name?, offset, limit) -> rows

Marti-AI's design choices (RE: dopis 1.5.2026 13:09):
  - Datum/čas → ISO string ('2026-09-01T08:00:00')
  - Prázdné buňky → null (None)
  - Čísla → vždy float (5.5)
  - Vzorce → computed value (openpyxl default data_only=True)
  - Errors (#N/A, #REF!) → null + entry v warnings list
  - Multi-sheet → Varianta C (separate list_excel_sheets + targeted reads)
  - Pagination → offset/limit, cap 500 rows per call

Implementační poznámky:
  - openpyxl je tranzitivně přes markitdown[all]; přidaný explicit
    do pyproject.toml pro defense-in-depth (gotcha #21 z 28.4.2026 --
    markdown lib také byla tranzitivní a v cloud venv chyběla).
  - data_only=True znamená, že openpyxl vrací CACHED computed values
    z xlsx. Pokud byl soubor upraven externě (např. otevřen v Numbers
    nebo neotevřen v Excelu po změně), vzorce mohou vracet None.
    V takovém případě warning: "vzorec bez cached value".
  - read_only=True optimalizuje memory/speed pro velké soubory; nevidí
    merged cells correctly, ale pro flat tabulky je to OK.
"""
from __future__ import annotations

import datetime as _dt
from pathlib import Path
from typing import Any

from core.logging import get_logger

logger = get_logger("rag.excel")

# Cap pro safety -- Marti-AI's volba (RE: dopis 1.5.2026)
MAX_ROWS_PER_CALL = 500
MAX_HEADERS_PREVIEW = 8  # v list_excel_sheets metadata


def _normalize_cell_value(value: Any, errors_collector: list[str], cell_ref: str = "") -> Any:
    """
    Normalize cell value per Marti-AI's design choices.

    - datetime/date/time → ISO string
    - errors (string starting with '#') → None + warning
    - int → float
    - everything else → as-is
    """
    if value is None:
        return None

    # Excel error markers vrací openpyxl jako string '#REF!', '#N/A', etc.
    if isinstance(value, str) and value.startswith("#") and value.endswith("!"):
        if cell_ref:
            errors_collector.append(f"{cell_ref}: {value}")
        else:
            errors_collector.append(value)
        return None
    if isinstance(value, str) and value in ("#N/A", "#NAME?", "#NULL!", "#DIV/0!", "#VALUE!", "#REF!", "#NUM!"):
        if cell_ref:
            errors_collector.append(f"{cell_ref}: {value}")
        else:
            errors_collector.append(value)
        return None

    if isinstance(value, (_dt.datetime, _dt.date, _dt.time)):
        try:
            return value.isoformat()
        except Exception:
            return str(value)

    if isinstance(value, bool):
        return value  # bool před int check (bool je subclass of int)

    if isinstance(value, int):
        return float(value)

    if isinstance(value, float):
        return value

    if isinstance(value, str):
        return value

    # Cokoli jiného (Decimal, custom types) -> string fallback
    return str(value)


def _resolve_document_path(document_id: int) -> tuple[str, str, int | None, int | None]:
    """
    Vrátí (storage_path, original_filename, tenant_id, project_id) pro document_id.

    Raises ValueError pokud document neexistuje, není xlsx, nebo nemá storage_path.
    """
    from core.database import get_session
    from modules.core.infrastructure.models_data import Document

    session = get_session()
    try:
        doc = session.query(Document).filter_by(id=document_id).first()
        if not doc:
            raise ValueError(f"Document id={document_id} nenalezen.")
        if not doc.storage_path:
            raise ValueError(f"Document id={document_id} nema storage_path.")

        # Validace file_type / extension
        ext = (doc.file_type or "").lower().lstrip(".")
        if ext not in ("xlsx", "xlsm"):
            raise ValueError(
                f"Document id={document_id} neni Excel (.xlsx/.xlsm), "
                f"file_type='{doc.file_type}'. Pro starsi .xls format "
                "prosim ulozit jako .xlsx (openpyxl nepodporuje legacy .xls)."
            )

        return (
            doc.storage_path,
            doc.original_filename or doc.name or f"document_{document_id}",
            doc.tenant_id,
            doc.project_id,
        )
    finally:
        session.close()


def _check_tenant_access(document_tenant_id: int | None, caller_tenant_id: int | None, is_parent: bool) -> bool:
    """
    Tenant scope check. Parent (is_marti_parent=True) bypass jako jinde.
    Document bez tenant_id (NULL) = inbox/global, accessible.
    """
    if is_parent:
        return True
    if document_tenant_id is None:
        return True  # inbox / global
    if caller_tenant_id is None:
        return False
    return document_tenant_id == caller_tenant_id


def list_excel_sheets(
    document_id: int,
    *,
    caller_tenant_id: int | None = None,
    is_parent: bool = False,
) -> dict:
    """
    Vrátí metadata pro všechny listy v xlsx.

    Output:
        {
            "filename": "rozvrh.xlsx",
            "n_sheets": 4,
            "sheets": [
                {
                    "name": "List1",
                    "n_rows": 234,
                    "n_cols": 8,
                    "headers_preview": ["Třída", "Učitel", "Předmět", ...]
                },
                ...
            ]
        }
    """
    try:
        import openpyxl
    except ImportError as e:
        raise RuntimeError(f"openpyxl neni dostupna: {e}")

    storage_path, filename, doc_tenant_id, _ = _resolve_document_path(document_id)

    if not _check_tenant_access(doc_tenant_id, caller_tenant_id, is_parent):
        raise PermissionError(
            f"Document id={document_id} patri jinemu tenantu (doc.tenant_id={doc_tenant_id})."
        )

    p = Path(storage_path)
    if not p.is_file():
        raise ValueError(f"Soubor neexistuje na disku: {storage_path}")

    wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
    try:
        sheets_info = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            n_rows = ws.max_row or 0
            n_cols = ws.max_column or 0

            # Headers preview = první row, max MAX_HEADERS_PREVIEW kolonek
            headers_preview: list[str] = []
            if n_rows >= 1:
                first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
                if first_row:
                    for cell in first_row[:MAX_HEADERS_PREVIEW]:
                        if cell is None:
                            headers_preview.append("")
                        else:
                            headers_preview.append(str(cell))

            sheets_info.append({
                "name": sheet_name,
                "n_rows": int(n_rows),
                "n_cols": int(n_cols),
                "headers_preview": headers_preview,
            })

        return {
            "document_id": document_id,
            "filename": filename,
            "n_sheets": len(sheets_info),
            "sheets": sheets_info,
        }
    finally:
        wb.close()


def read_excel_structured(
    document_id: int,
    *,
    sheet_name: str | None = None,
    sheet_index: int | None = None,
    offset: int = 0,
    limit: int = MAX_ROWS_PER_CALL,
    caller_tenant_id: int | None = None,
    is_parent: bool = False,
) -> dict:
    """
    Vrátí structured rows z konkrétního listu.

    Output:
        {
            "document_id": 42,
            "filename": "rozvrh.xlsx",
            "sheet": "Učitelé",
            "headers": ["Jméno", "Aprobace", "Úvazek"],
            "rows": [
                {"Jméno": "Nováková", "Aprobace": "M, F", "Úvazek": 1.0},
                ...
            ],
            "n_rows_total": 23,       # celkem rows v listu (mimo header)
            "offset": 0,
            "limit": 500,
            "returned": 23,            # kolik vraceno v tomto callu
            "has_more": False,
            "warnings": []             # napr. "B7: #REF!"
        }
    """
    try:
        import openpyxl
    except ImportError as e:
        raise RuntimeError(f"openpyxl neni dostupna: {e}")

    # Validace pagination
    if offset < 0:
        raise ValueError("offset musi byt >= 0")
    if limit < 1:
        raise ValueError("limit musi byt >= 1")
    if limit > MAX_ROWS_PER_CALL:
        limit = MAX_ROWS_PER_CALL  # tichá clamp

    if sheet_name is not None and sheet_index is not None:
        raise ValueError("Pouze jeden z 'sheet_name' nebo 'sheet_index' (ne oba).")

    storage_path, filename, doc_tenant_id, _ = _resolve_document_path(document_id)

    if not _check_tenant_access(doc_tenant_id, caller_tenant_id, is_parent):
        raise PermissionError(
            f"Document id={document_id} patri jinemu tenantu (doc.tenant_id={doc_tenant_id})."
        )

    p = Path(storage_path)
    if not p.is_file():
        raise ValueError(f"Soubor neexistuje na disku: {storage_path}")

    wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
    try:
        # Resolve sheet
        if sheet_name is not None:
            if sheet_name not in wb.sheetnames:
                raise ValueError(
                    f"Sheet '{sheet_name}' neexistuje. Dostupne: {wb.sheetnames}"
                )
            ws = wb[sheet_name]
            resolved_name = sheet_name
        elif sheet_index is not None:
            if sheet_index < 0 or sheet_index >= len(wb.sheetnames):
                raise ValueError(
                    f"sheet_index={sheet_index} mimo rozsah (0..{len(wb.sheetnames)-1})."
                )
            resolved_name = wb.sheetnames[sheet_index]
            ws = wb[resolved_name]
        else:
            # Default: prvni sheet
            resolved_name = wb.sheetnames[0]
            ws = wb[resolved_name]

        n_rows_max = ws.max_row or 0
        n_cols_max = ws.max_column or 0

        warnings: list[str] = []

        # Read headers (first row)
        headers: list[str] = []
        if n_rows_max >= 1:
            first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if first_row:
                for idx, cell in enumerate(first_row):
                    if cell is None:
                        headers.append(f"_col_{idx + 1}")  # placeholder pro prazdny header
                    else:
                        headers.append(str(cell))

        n_data_rows_total = max(0, n_rows_max - 1)  # bez header row

        # Pagination: data rows are 2..n_rows_max
        # offset 0 = start od row 2
        # offset 5 = start od row 7
        start_row = 2 + offset
        end_row = min(n_rows_max, start_row + limit - 1)

        rows_out: list[dict] = []
        if start_row <= n_rows_max and headers:
            for row_idx, row_values in enumerate(
                ws.iter_rows(min_row=start_row, max_row=end_row, values_only=True),
                start=start_row,
            ):
                # iter_rows muze vratit kratsi tuple nez headers; padding
                row_dict = {}
                for col_idx, header in enumerate(headers):
                    if col_idx < len(row_values):
                        raw_val = row_values[col_idx]
                    else:
                        raw_val = None
                    cell_ref = f"{_col_letter(col_idx + 1)}{row_idx}"
                    row_dict[header] = _normalize_cell_value(raw_val, warnings, cell_ref)
                rows_out.append(row_dict)

        returned = len(rows_out)
        has_more = (offset + returned) < n_data_rows_total

        return {
            "document_id": document_id,
            "filename": filename,
            "sheet": resolved_name,
            "headers": headers,
            "rows": rows_out,
            "n_rows_total": n_data_rows_total,
            "offset": offset,
            "limit": limit,
            "returned": returned,
            "has_more": has_more,
            "warnings": warnings,
        }
    finally:
        wb.close()


def _col_letter(col_idx: int) -> str:
    """1 -> 'A', 26 -> 'Z', 27 -> 'AA', etc."""
    result = ""
    while col_idx > 0:
        col_idx, rem = divmod(col_idx - 1, 26)
        result = chr(ord("A") + rem) + result
    return result
