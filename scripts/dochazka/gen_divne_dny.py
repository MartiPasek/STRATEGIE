# -*- coding: utf-8 -*-
"""
DIVNÉ DNY V DOCHÁZCE — přehled do .xlsx pro jednoho člověka
Claude-26 / Peťa, 3. 8. 2026

K čemu to je
------------
Na rozdíl od kontrolní tabulky (ta ukazuje celý měsíc) tenhle přehled vypíše
JEN dny, kde je něco podezřelého. Prázdný přehled = všechno v pořádku.

Jak se to pouští (dva kroky)
----------------------------
1) Data z databáze přes SQL most:
   - vezmi scripts/dochazka/divne_dny.sql
   - nahraď {CISLO_ZAM}, {OD}, {DO}
   - vlož do scripts/claude_sql/CLAUDE2_SQL.sql a spusť CLAUDE2_GO.txt (db=pg)
   - výsledek si ulož z scripts/claude_sql/CLAUDE2_OUT_FULL.txt
2) Tenhle skript:
   python3 gen_divne_dny.py data.tsv Divne_dny_07_Artim.xlsx "Josef Artim" 498 "červenec 2026"

Meze (zadala Peťa 3. 8. 2026)
-----------------------------
* moc dlouhý den = víc než 10 h
* moc krátký den = míň než 6 h (a není to absence)
* dlouhá pauza = přes hodinu
* překryv = dvě práce na sobě; odhlášení „Dnes už se mnou nepočítej" se
  ZÁMĚRNĚ ignoruje — lidé jinou možnost v aplikaci nemají, takže jeho přesah
  do odpoledne je normální stav, ne chyba.
"""
import sys
import csv

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

C_HLAV = "1C2636"
C_ZELENA = "E2EFDA"
C_ZELENA_TEXT = "1E6B3A"
C_CERVENA = "FBD5D5"
C_CERVENA_TEXT = "9C1616"
C_ORANZ = "FCE4D6"
C_ORANZ_TEXT = "833C0B"
C_MODRA = "DCE6F1"
C_MODRA_TEXT = "1F4E79"

# jak se který nález obarví — vážné věci červeně, spíš upozornění oranžově
BARVY = {
    "Chybí zápis": (C_CERVENA, C_CERVENA_TEXT),
    "Docházka a rozpad nesedí": (C_CERVENA, C_CERVENA_TEXT),
    "Chybí zakázka nebo činnost": (C_CERVENA, C_CERVENA_TEXT),
    "Překryv": (C_CERVENA, C_CERVENA_TEXT),
    "Neukončený den": (C_ORANZ, C_ORANZ_TEXT),
    "Moc dlouhý den": (C_ORANZ, C_ORANZ_TEXT),
    "Moc krátký den": (C_ORANZ, C_ORANZ_TEXT),
    "Dlouhá pauza": (C_MODRA, C_MODRA_TEXT),
}

SLOUPCE = [("Datum", 14), ("Den", 7), ("Co je divně", 26),
           ("Podrobnost", 52), ("Hodiny", 10)]

MERITKA = [
    "Moc dlouhý den = víc než 10 hodin",
    "Moc krátký den = míň než 6 hodin (a není to absence)",
    "Dlouhá pauza = přestávka přes hodinu",
    "Překryv = dvě práce na sobě (odhlášení „Dnes už se mnou nepočítej\" se nepočítá,",
    "        lidé jinou možnost v aplikaci nemají)",
    "Chybí zápis = pracovní den, o kterém systém neví vůbec nic",
    "Docházka a rozpad nesedí = součty dne se liší o víc než 0,05 h",
]


def postav(vstup_tsv, vystup_xlsx, jmeno=None, cislo=None, obdobi=""):
    with open(vstup_tsv, encoding="utf-8") as f:
        radky = list(csv.DictReader(f, delimiter="\t"))

    if radky:
        jmeno = jmeno or radky[0].get("jmeno") or ""
        cislo = cislo or radky[0].get("cislo_zam") or ""

    wb = Workbook()
    ws = wb.active
    ws.title = "Divné dny"

    nadpis = "Divné dny v docházce — %s%s%s" % (
        jmeno or "?", (" (č. %s)" % cislo) if cislo else "",
        (", %s" % obdobi) if obdobi else "")
    c = ws.cell(1, 1, nadpis)
    c.font = Font(bold=True, size=14)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(SLOUPCE))
    ws.row_dimensions[1].height = 22

    for i, (nazev, sirka) in enumerate(SLOUPCE, start=1):
        b = ws.cell(3, i, nazev)
        b.fill = PatternFill("solid", fgColor=C_HLAV)
        b.font = Font(color="FFFFFF", bold=True, size=11)
        ws.column_dimensions[get_column_letter(i)].width = sirka
    ws.freeze_panes = "A4"

    r = 4
    if not radky:
        c = ws.cell(r, 1, "✓ Všechno v pořádku — nenašel jsem jediný podezřelý den.")
        c.font = Font(bold=True, size=12, color=C_ZELENA_TEXT)
        for i in range(1, len(SLOUPCE) + 1):
            ws.cell(r, i).fill = PatternFill("solid", fgColor=C_ZELENA)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(SLOUPCE))
        r += 2
    else:
        for z in radky:
            problem = z.get("problem") or ""
            pozadi, pismo = BARVY.get(problem, (C_ORANZ, C_ORANZ_TEXT))
            hodnoty = [z.get("datum"), z.get("den_v_tydnu"), problem,
                       z.get("detail"), z.get("hodiny")]
            for i, v in enumerate(hodnoty, start=1):
                cc = ws.cell(r, i, v or None)
                cc.fill = PatternFill("solid", fgColor=pozadi)
                cc.font = Font(color=pismo, bold=(i == 3), size=11)
                if i == 5:
                    cc.alignment = Alignment(horizontal="right")
            r += 1
        r += 1
        c = ws.cell(r, 1, "Celkem podezřelých řádků: %d" % len(radky))
        c.font = Font(bold=True, size=11)
        r += 2

    c = ws.cell(r, 1, "Podle čeho se to posuzuje")
    c.font = Font(bold=True, size=11)
    r += 1
    for m in MERITKA:
        ws.cell(r, 1, m).font = Font(size=10, color="5A6675")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(SLOUPCE))
        r += 1

    wb.save(vystup_xlsx)
    return {"soubor": vystup_xlsx, "nalezu": len(radky), "jmeno": jmeno}


if __name__ == "__main__":
    if len(sys.argv) < 3:
        raise SystemExit(__doc__)
    vysledek = postav(sys.argv[1], sys.argv[2],
                      sys.argv[3] if len(sys.argv) > 3 else None,
                      sys.argv[4] if len(sys.argv) > 4 else None,
                      sys.argv[5] if len(sys.argv) > 5 else "")
    print("Hotovo: %(soubor)s · %(jmeno)s · nálezů: %(nalezu)s" % vysledek)
