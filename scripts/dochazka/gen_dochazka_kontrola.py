# -*- coding: utf-8 -*-
"""
KONTROLNÍ TABULKA DOCHÁZKY PRO JEDNOHO ČLOVĚKA (.xlsx)
Claude-26 / Peťa, 3. 8. 2026

K čemu to je
------------
Člověk dostane mailem svůj měsíc: vlevo zamčená data ze systému, vpravo žluté
kolonky, do kterých dopíše opravy. Vrátí soubor a my podle něj docházku srovnáme.

Jak se to pouští (dva kroky)
----------------------------
1) Data z databáze přes SQL most:
   - vezmi scripts/dochazka/dochazka_kontrola_data.sql
   - nahraď {CISLO_ZAM}, {OD}, {DO}
   - vlož do scripts/claude_sql/CLAUDE2_SQL.sql a spusť CLAUDE2_GO.txt (db=pg)
   - výsledek si ulož z scripts/claude_sql/CLAUDE2_OUT_FULL.txt
2) Tenhle skript:
   python3 gen_dochazka_kontrola.py data.tsv Dochazka_07_Kolarova.xlsx

Co tabulka umí (a proč)
-----------------------
* Jeden soubor = jeden člověk. List je zamčený BEZ hesla, psát jde jen do žlutých
  sloupců vpravo.
* Šedý řádek se součtem za každý den, dole součet za celé období.
* Přestávky zeleně kurzívou, do součtu se NEpočítají.
* Absence (dovolená, lékař, nemoc) MUSÍ být ve vstupu — jinak by se dovolená
  ukázala jako "chybí zápis". (Na tom se Claude-26 spálil 31. 7. 2026.)
* Chybějící pracovní den = červený řádek "⚠ CHYBÍ ZÁPIS" (jen do včerejška).
* NOVÉ (Peťa 3. 8. 2026): když docházka zapsaná JE, ale chybí u ní zakázka
  a činnost, řádek se ukáže i s časy a červeně se vyznačí jen to, co chybí,
  s výzvou, ať nám to člověk doplní. Dřív takový den vypadal jako úplně prázdný
  a lidi zbytečně vypisovali časy, které v systému dávno byly.
* NOVÉ (Peťa 3. 8. 2026): sloupce "pauza od" / "pauza do", aby šlo přestávku
  DOPLNIT i tam, kde v systému žádná zapsaná není. Z časů od–do pak umíme
  správně rozdělit práci před pauzou a po ní.
* Víkend modře, svátek oranžově s názvem.

ID záznamu je SLOŽENÉ a jinak není jednoznačné — vždy párovat na obě části:
  W-<id> úsek rozpadu na zakázky · P-/A-<id> docházkový záznam / absence
  B-<id> přestávka · C-<id> denní souhrn
"""
import sys
import csv
import datetime as dt

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Protection
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------- vzhled
C_HLAV_POZADI = "1C2636"   # tmavá hlavička levé části
C_ZLUTA_HLAV = "BF8F00"    # hlavička opravných sloupců
C_ZLUTA = "FFF2CC"         # žluté kolonky k vyplnění
C_SOUCET = "EDEFF2"        # šedý řádek se součtem dne
C_PAUZA = "E2EFDA"         # zelený řádek přestávky
C_PAUZA_TEXT = "5A6675"
C_CERVENA = "FBD5D5"       # chybí zápis / chybí zakázka a činnost
C_CERVENA_TEXT = "9C1616"
C_VIKEND = "DCE6F1"
C_VIKEND_TEXT = "1F4E79"
C_SVATEK = "FCE4D6"
C_SVATEK_TEXT = "833C0B"

DNY = ["Po", "Ut", "St", "Ct", "Pa", "So", "Ne"]

SLOUPCE = [
    ("ID záznamu", 12.5), ("Zakázka", 13.5), ("Jméno a příjmení", 23.5),
    ("Číslo zam.", 10.5), ("Druh činnosti", 12.5), ("Činnost", 25.5),
    ("Den", 7.5), ("Začátek", 18.5), ("Konec", 18.5), ("Hodiny", 10.5),
    ("Poznámka", 19.5), ("", 2.5),
]
OPRAVY = [
    ("oprava zakázka\n(číslo zakázky)", 17.5),
    ("oprava druh činnosti\n(číslo)", 17.5),
    ("oprava činnost\n(název)", 21.5),
    ("oprava začátek\n(čas, např. 7:30)", 18.5),
    ("oprava konec\n(čas, např. 16:00)", 18.5),
    ("pauza od\n(čas, např. 12:00)", 18.5),
    ("pauza do\n(čas, např. 12:30)", 18.5),
    ("smazat řádek\n(napište x)", 14.5),
    ("poznámka k opravě", 27.5),
]
PRVNI_OPRAVNY = len(SLOUPCE) + 1          # sloupec M
POSLEDNI = len(SLOUPCE) + len(OPRAVY)     # sloupec U

VYZVA = ("Nevíme, na čem jste tenhle den dělal(a) — napište nám to prosím "
         "do žlutých sloupců vpravo.")


# ---------------------------------------------------------------- pomůcky
def cti_tsv(cesta):
    """Načte výstup z SQL mostu (CLAUDE2_OUT_FULL.txt) jako seznam slovníků."""
    with open(cesta, encoding="utf-8") as f:
        return list(csv.DictReader(f, delimiter="\t"))


def na_datum(s):
    return dt.date.fromisoformat(s[:10]) if s else None


def na_cas(s):
    """'2026-07-01T08:28:00+02:00' → datetime bez časové zóny (lokální čas)."""
    if not s:
        return None
    return dt.datetime.fromisoformat(s).replace(tzinfo=None)


def na_cislo(s):
    try:
        return round(float(s), 2)
    except (TypeError, ValueError):
        return None


def hodiny_a_minuty(h):
    """0.75 → '0:45'"""
    celkem = int(round((h or 0) * 60))
    return "%d:%02d" % (celkem // 60, celkem % 60)


def dtm(x):
    return x.strftime("%d.%m.%Y %H:%M") if x else ""


# ---------------------------------------------------------------- stavba
class Tabulka:
    def __init__(self, ws):
        self.ws = ws
        self.r = 1

    def hlavicka(self):
        ws = self.ws
        for i, (nazev, sirka) in enumerate(SLOUPCE + OPRAVY, start=1):
            c = ws.cell(1, i)
            c.value = nazev or None
            ws.column_dimensions[get_column_letter(i)].width = sirka
            if i == len(SLOUPCE):          # prázdný oddělovací sloupec
                continue
            if i < len(SLOUPCE):
                c.fill = PatternFill("solid", fgColor=C_HLAV_POZADI)
                c.font = Font(color="FFFFFF", bold=True, size=11)
            else:
                c.fill = PatternFill("solid", fgColor=C_ZLUTA_HLAV)
                c.font = Font(color="FFFFFF", bold=True, size=10)
                c.alignment = Alignment(wrap_text=True, vertical="center")
        ws.row_dimensions[1].height = 30
        ws.freeze_panes = "A2"
        self.r = 2

    def _zlute(self, radek, pozadi=None):
        """Odemkne opravné sloupce, ať do nich jde psát i v zamčeném listu."""
        for i in range(PRVNI_OPRAVNY, POSLEDNI + 1):
            c = self.ws.cell(radek, i)
            c.fill = PatternFill("solid", fgColor=pozadi or C_ZLUTA)
            c.protection = Protection(locked=False)

    def _obarvi(self, radek, pozadi, barva_pisma, tucne=False, kurziva=False):
        for i in range(1, len(SLOUPCE)):
            c = self.ws.cell(radek, i)
            if pozadi:
                c.fill = PatternFill("solid", fgColor=pozadi)
            c.font = Font(color=barva_pisma, bold=tucne, italic=kurziva, size=12)

    def radek_zaznamu(self, z, jmeno, cislo_zam, den_typ):
        """Jeden datový řádek (úsek práce, přestávka, absence nebo chybějící rozpad)."""
        ws, r = self.ws, self.r
        chybi_rozpad = (z["druh"] == "P")
        je_pauza = (z["druh"] == "B")

        ws.cell(r, 1, z["id"])
        ws.cell(r, 2, z["zakazka"] or None)
        ws.cell(r, 3, jmeno)
        ws.cell(r, 4, cislo_zam)
        ws.cell(r, 5, int(z["ec_cislo"]) if z["ec_cislo"] else None)
        ws.cell(r, 6, z["cinnost"] or None)
        ws.cell(r, 7, DNY[z["_datum"].weekday()])
        ws.cell(r, 8, dtm(z["_od"]))
        ws.cell(r, 9, dtm(z["_konec"]))
        c_hod = ws.cell(r, 10, z["_hodiny"] if z["_hodiny"] is not None else 0)
        c_hod.number_format = "0.00"
        ws.cell(r, 11, z["poznamka"] or None)

        # barva podle druhu dne / řádku
        if je_pauza:
            self._obarvi(r, C_PAUZA, C_PAUZA_TEXT, kurziva=True)
        elif den_typ == "svatek":
            self._obarvi(r, C_SVATEK, C_SVATEK_TEXT)
        elif den_typ == "vikend":
            self._obarvi(r, C_VIKEND, C_VIKEND_TEXT)
        else:
            self._obarvi(r, None, "000000")

        # NOVÉ: chybí zakázka a činnost → červeně jen ty kolonky + výzva
        if chybi_rozpad:
            for i in (2, 5, 6):
                c = ws.cell(r, i)
                c.fill = PatternFill("solid", fgColor=C_CERVENA)
                c.font = Font(color=C_CERVENA_TEXT, bold=True, size=12)
            if not z["zakazka"]:
                ws.cell(r, 2, "⚠ doplňte")
            ws.cell(r, 6, "⚠ doplňte činnost")
            c = ws.cell(r, 11, VYZVA)
            c.fill = PatternFill("solid", fgColor=C_CERVENA)
            c.font = Font(color=C_CERVENA_TEXT, bold=True, size=11)
            c.alignment = Alignment(wrap_text=True, vertical="center")

        self._zlute(r)
        self.r += 1

    def radek_chybi(self, den, jmeno, cislo_zam):
        ws, r = self.ws, self.r
        ws.cell(r, 3, jmeno)
        ws.cell(r, 4, cislo_zam)
        ws.cell(r, 6, "⚠ CHYBÍ ZÁPIS")
        ws.cell(r, 7, DNY[den.weekday()])
        ws.cell(r, 8, den.strftime("%d.%m.%Y"))
        ws.cell(r, 11, VYZVA)
        self._obarvi(r, C_CERVENA, C_CERVENA_TEXT, tucne=True)
        ws.cell(r, 11).alignment = Alignment(wrap_text=True, vertical="center")
        self._zlute(r, C_CERVENA)
        self.r += 1

    def radek_souctu(self, den, prace, pauzy, svatek_nazev=None):
        ws, r = self.ws, self.r
        popis = "součet za " + den.strftime("%d.%m.%Y")
        if pauzy > 0.004:
            popis += " · pauzy " + hodiny_a_minuty(pauzy)
        if svatek_nazev:
            popis += " · svátek " + svatek_nazev
        for i in range(1, POSLEDNI + 1):
            if i == len(SLOUPCE):
                continue
            ws.cell(r, i).fill = PatternFill("solid", fgColor=C_SOUCET)
        c = ws.cell(r, 9, popis)
        c.font = Font(color=C_PAUZA_TEXT, italic=True, size=10)
        c.alignment = Alignment(horizontal="right")
        c = ws.cell(r, 10, round(prace, 2))
        c.font = Font(bold=True, size=12)
        c.number_format = "0.00"
        self.r += 1

    def zavery(self, celkem):
        ws = self.ws
        self.r += 1
        c = ws.cell(self.r, 9, "celkem za období")
        c.font = Font(bold=True, size=12)
        c.alignment = Alignment(horizontal="right")
        c = ws.cell(self.r, 10, round(celkem, 2))
        c.font = Font(bold=True, size=12)
        c.number_format = "0.00"
        self.r += 2
        ws.cell(self.r, 2, "Chybí-li ještě něco, dopište sem:").font = Font(bold=True)
        self.r += 1
        for _ in range(8):                       # prázdné řádky na dopsání
            self._zlute(self.r)
            self.r += 1


def postav(vstup_tsv, vystup_xlsx):
    data = cti_tsv(vstup_tsv)

    jmeno, cislo_zam = "", ""
    kalendar = {}          # den -> (typ, nazev_svatku)
    podle_dne = {}         # den -> [zaznamy]

    for z in data:
        druh = z["druh"]
        if druh == "OSOBA":
            jmeno, cislo_zam = z["cinnost"], z["poznamka"]
            continue
        if druh == "ZAMEK":
            continue
        den = na_datum(z["datum"])
        if den is None:
            continue
        if druh == "DEN":
            if z["ec_cislo"] == "1":
                kalendar[den] = ("svatek", z["cinnost"] or None)
            elif z["zakazka"] == "pracovni":
                kalendar[den] = ("pracovni", None)
            else:
                kalendar[den] = ("vikend", None)
            continue
        z["_datum"] = den
        z["_od"] = na_cas(z["od"])
        z["_konec"] = na_cas(z["konec"])
        z["_hodiny"] = na_cislo(z["hodiny"]) or 0.0
        podle_dne.setdefault(den, []).append(z)

    if not jmeno:
        raise SystemExit("Ve vstupu chybí řádek OSOBA — zkontroluj, že SQL našlo "
                         "člověka podle osobního čísla.")

    wb = Workbook()
    ws = wb.active
    ws.title = "Docházka"
    t = Tabulka(ws)
    t.hlavicka()

    dnes = dt.date.today()
    vsechny_dny = sorted(set(list(kalendar) + list(podle_dne)))
    celkem = 0.0

    for den in vsechny_dny:
        typ, svatek = kalendar.get(den, ("pracovni", None))
        zaznamy = podle_dne.get(den, [])

        if not zaznamy:
            # prázdný PRACOVNÍ den do včerejška = chybí zápis; víkend/svátek mlčí
            if typ == "pracovni" and den < dnes:
                t.radek_chybi(den, jmeno, cislo_zam)
            continue

        zaznamy.sort(key=lambda z: (z["_od"] or dt.datetime.min, z["id"]))
        prace = pauzy = 0.0
        for z in zaznamy:
            t.radek_zaznamu(z, jmeno, cislo_zam, typ)
            if z["druh"] == "B":
                pauzy += z["_hodiny"]
            else:
                prace += z["_hodiny"]
        t.radek_souctu(den, prace, pauzy, svatek)
        celkem += prace

    t.zavery(celkem)
    ws.protection.sheet = True          # zamčeno BEZ hesla (heslo se nenastavuje)
    wb.save(vystup_xlsx)
    return {"soubor": vystup_xlsx, "jmeno": jmeno, "cislo_zam": cislo_zam,
            "dnu": len(podle_dne), "celkem": round(celkem, 2)}


if __name__ == "__main__":
    if len(sys.argv) < 3:
        raise SystemExit(__doc__)
    vysledek = postav(sys.argv[1], sys.argv[2])
    print("Hotovo: %(soubor)s · %(jmeno)s (č. %(cislo_zam)s) · "
          "%(dnu)s dnů · celkem %(celkem)s h" % vysledek)
