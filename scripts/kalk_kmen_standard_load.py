#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
kalk_kmen_standard_load.py — opakovatelný loader STANDARD kalkulace (Excel) -> proj.kalk_kmen

Autor: Claude-24 (Kristý). 23.7.2026 zal., 24.7.2026 rozšířeno (dedup dle priority listů,
sloupce cena_cc/nc/rabatt, idempotentní append).

CO DĚLÁ
  Projde všechny listy STANDARD sešitu s hlavičkou "Typ / Bestell.-Nr.", vezme řádky
  s vyplněným obj. číslem (C) a výrobcem (D), SLOUČÍ duplicity na jeden díl = jeden řádek
  (dle priority listů) a vygeneruje SQL pro proj.kalk_kmen.
    - default    : idempotentní APPEND — INSERT ... ON CONFLICT (reg_cis) DO NOTHING
                   (dohraje jen díly, které v katalogu ještě nejsou; katalog roste).
    - --rebuild  : čistý REBUILD — DELETE FROM proj.kalk_kmen; INSERT ... (plná náhrada).

MAPOVÁNÍ SLOUPCŮ (pevně dle Kristý — bere se podle POZICE písmene, ne podle labelu):
  A (1)  Pos.               -> excel_pozice   (integer)
  B (2)  Bezeichnung        -> oznaceni
  C (3)  Typ / Bestell.-Nr. -> reg_cis        (identita dílu, UNIQUE)
  D (4)  Lieferant          -> vyrobce
  F (6)  Einheitspreis      -> cena_cc_ref    (ceníková cena)
  H (8)  Koeffiz.           -> k_arb A ZÁROVEŇ k_vkm  (stejná hodnota)
  I (9)  Bemerkung          -> rabatt_ref     (procenta vyextrahovaná z textu, např. "74%; Federzug" -> 74)
  J (10) Hmotnost           -> hmotnost_kg
  M (13) Einheitpreis       -> cena_nc_ref    (počítaná jednotková cena vč. rabattu; může být > cena_cc)
  navíc: excel_soubor = název souboru, excel_list = list, excel_radek = číslo řádku,
         zdroj = 'STANDARD_xlsm', synced_at = now(). nazev se NEPLNÍ.

DEDUP (jeden díl = jeden řádek, klíč = reg_cis)
  Priorita listů: Základní 2026 > Rozšíření 2025 > ICOTEK 2025 > Šínový systém > UL508A
                  > Sie Schraube > Siemens 3SU1+ Metall.
  Když je reg_cis na víc listech, vyhraje list s vyšší prioritou. V rámci téhož listu
  (díl vícekrát) vyhraje řádek s koeficientem, pak s cenou, pak nejnižší excel_radek.

FILTR ŘÁDKŮ
  obj. číslo (C) neprázdné + obsahuje číslici + není hlavičkový zbytek (., SLAVE ID, MAT ID);
  výrobce (D) neprázdný a != '.' (tím vypadnou poznámky/oddílové řádky).

POUŽITÍ
  python3 kalk_kmen_standard_load.py <STANDARD.xlsm> [--out CLAUDE_SQL.sql] [--rebuild]
  Tip: --out scripts/claude_sql/CLAUDE_SQL.sql, pak přes most (db=pg) = schvalovací banner.

POZN.: Schéma proj.kalk_kmen musí mít sloupce (id PK, vyrobce, oznaceni, k_arb, k_vkm,
       hmotnost_kg, cena_cc_ref, cena_nc_ref, rabatt_ref, excel_*), kmen_ec_id nullable,
       UNIQUE index na reg_cis. Založeno 23.-24.7.2026 (req #1365/#1370/#1394). Loader schéma nemění.
"""
import sys, os, re, argparse
import openpyxl

SKIP_SHEETS = {"Poznámky", "ukladani_data", "Rešerše"}
JUNK_REG = {".", "SLAVE ID", "MAT ID", "Typ / Bestell.-Nr.", "Bestell.-Nr.", ""}

# priorita listů (nižší číslo = přednost)
SHEET_PRIORITY = {
    "Základní 2026": 1, "Rozšíření 2025": 2, "ICOTEK 2025": 3,
    "Šínový systém": 4, "UL508A": 5, "Sie Schraube": 6, "Siemens 3SU1+ Metall": 7,
}

# 0-indexované pozice sloupců
COL_POZICE, COL_OZN, COL_REG, COL_VYR = 0, 1, 2, 3      # A, B, C, D
COL_CC, COL_KOEF, COL_BEMERK, COL_HMOT, COL_NC = 5, 7, 8, 9, 12  # F, H, I, J, M


def norm(v):
    return ("" if v is None else str(v)).strip()

def has_digit(s):
    return any(ch.isdigit() for ch in s)

def find_header_row(ws):
    for i, r in enumerate(ws.iter_rows(min_row=1, max_row=12, values_only=True), 1):
        if "Bestell" in " | ".join(norm(x) for x in r):
            return i
    return None

def fnum(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).strip().replace(",", "."))
    except ValueError:
        return None

def rabat_from_text(v):
    """Vyextrahuj úvodní procenta z textu typu '74%; Federzug' -> 74.0. Jinak None."""
    if v is None:
        return None
    m = re.match(r'\s*(\d+(?:[.,]\d+)?)\s*%', str(v))
    return float(m.group(1).replace(",", ".")) if m else None

def cell(row, idx):
    return row[idx] if len(row) > idx else None

def parse_rows(xlsm_path):
    fn = os.path.basename(xlsm_path)
    wb = openpyxl.load_workbook(xlsm_path, data_only=True, read_only=True)
    rows = []
    for ws in wb.worksheets:
        if ws.title in SKIP_SHEETS:
            continue
        hi = find_header_row(ws)
        if hi is None:
            continue
        for ri, r in enumerate(ws.iter_rows(min_row=hi + 1, values_only=True), start=hi + 1):
            reg = norm(cell(r, COL_REG))
            vyr = norm(cell(r, COL_VYR))
            if reg in JUNK_REG or not has_digit(reg):
                continue
            if not vyr or vyr == ".":
                continue
            rows.append(dict(
                reg_cis=reg, oznaceni=norm(cell(r, COL_OZN)) or None, vyrobce=vyr,
                hmotnost_kg=fnum(cell(r, COL_HMOT)), koef=fnum(cell(r, COL_KOEF)),
                cena_cc_ref=fnum(cell(r, COL_CC)), cena_nc_ref=fnum(cell(r, COL_NC)),
                rabatt_ref=rabat_from_text(cell(r, COL_BEMERK)),
                excel_soubor=fn, excel_list=ws.title, excel_radek=ri,
                excel_pozice=cell(r, COL_POZICE)))
    return rows

def dedup(rows):
    """Jeden díl = jeden řádek. Priorita listu, pak koef, pak cena, pak nejnižší řádek."""
    def rank(x):
        return (SHEET_PRIORITY.get(x["excel_list"], 99),
                0 if x["koef"] is not None else 1,
                0 if x["cena_cc_ref"] is not None else 1,
                x["excel_radek"])
    best = {}
    for x in rows:
        k = x["reg_cis"]
        if k not in best or rank(x) < rank(best[k]):
            best[k] = x
    return sorted(best.values(), key=lambda x: (SHEET_PRIORITY.get(x["excel_list"], 99), x["excel_radek"]))

def sql_q(s):
    return "NULL" if s is None else "'" + str(s).replace("'", "''") + "'"

def sql_num(v):
    return "NULL" if v is None else repr(float(v))

def sql_int(v):
    if v is None:
        return "NULL"
    try:
        return str(int(float(str(v).replace(",", "."))))
    except (ValueError, TypeError):
        return "NULL"

def build_sql(xlsm_path, rebuild=False):
    final = dedup(parse_rows(xlsm_path))
    vals = []
    for x in final:
        vals.append("(" + ",".join([
            sql_q(x["reg_cis"]), sql_q(x["oznaceni"]), sql_q(x["vyrobce"]),
            sql_num(x["hmotnost_kg"]), sql_num(x["koef"]), sql_num(x["koef"]),
            sql_num(x["cena_cc_ref"]), sql_num(x["cena_nc_ref"]), sql_num(x["rabatt_ref"]),
            sql_q(x["excel_soubor"]), sql_q(x["excel_list"]),
            sql_int(x["excel_radek"]), sql_int(x["excel_pozice"]),
            sql_q("STANDARD_xlsm"), "now()",
        ]) + ")")
    cols = ("reg_cis, oznaceni, vyrobce, hmotnost_kg, k_vkm, k_arb, "
            "cena_cc_ref, cena_nc_ref, rabatt_ref, "
            "excel_soubor, excel_list, excel_radek, excel_pozice, zdroj, synced_at")
    parts = []
    if rebuild:
        parts.append("DELETE FROM proj.kalk_kmen;")
    if vals:
        insert = "INSERT INTO proj.kalk_kmen (" + cols + ") VALUES\n" + ",\n".join(vals)
        if not rebuild:
            insert += "\nON CONFLICT (reg_cis) DO NOTHING"
        parts.append(insert + ";")
    return "\n".join(parts) + "\n", len(final)

def main():
    ap = argparse.ArgumentParser(description="STANDARD Excel -> proj.kalk_kmen (dedup, idempotentní)")
    ap.add_argument("xlsm", help="cesta ke STANDARD sešitu (.xlsm)")
    ap.add_argument("--out", help="soubor pro SQL (default stdout)")
    ap.add_argument("--rebuild", action="store_true",
                    help="čistý rebuild (DELETE + INSERT); default = idempotentní append (ON CONFLICT DO NOTHING)")
    args = ap.parse_args()
    if not os.path.exists(args.xlsm):
        sys.exit(f"CHYBA: soubor neexistuje: {args.xlsm}")

    sql, n = build_sql(args.xlsm, rebuild=args.rebuild)
    mode = "REBUILD (DELETE+INSERT)" if args.rebuild else "APPEND (ON CONFLICT DO NOTHING)"
    print(f"  režim: {mode} · dílů po dedup: {n}", file=sys.stderr)

    if args.out:
        tmp = args.out + ".tmp"
        with open(tmp, "w", newline="") as f:
            f.write(sql)
        os.replace(tmp, args.out)
        print(f"  SQL zapsáno do {args.out} ({len(sql)} B).", file=sys.stderr)
    else:
        sys.stdout.write(sql)

if __name__ == "__main__":
    main()
